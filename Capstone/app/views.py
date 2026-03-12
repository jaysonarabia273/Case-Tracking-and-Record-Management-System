from datetime import timezone
from unittest import case
from django.contrib.auth.models import User
from tokenize import Comment
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate
from .forms import UserSignUpForm, CommentForm, AppointmentForm
from .models import Notification, Profile, Case, CaseStatus
from django.utils.crypto import get_random_string
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from .models import Notification, GuidanceSession, Profile
from .models import Notification, EmailVerification
from django.core.mail import send_mail
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, FileResponse
from django.db.models import Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
import json
from .models import Notification, GuidanceSession
from django.views.decorators.csrf import csrf_exempt
from .models import Notification, StudentEvaluation, Profile, EvaluationReport
import logging
from django.views.decorators.http import require_http_methods
from .forms import GuidanceSessionForm
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_decode
from django.urls import reverse
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.conf import settings

logger = logging.getLogger(__name__)

@login_required
@require_http_methods(["POST"])
def update_session_status(request, session_id):
    try:
        # Get the session object
        session = get_object_or_404(GuidanceSession, id=session_id)
        
        # Parse request data
        data = json.loads(request.body)
        new_status = data.get('status')
        
        
        # Validate the status
        valid_statuses = [choice[0] for choice in session.STATUS_CHOICES]
        if new_status not in valid_statuses:
            return JsonResponse({'success': False, 'error': 'Invalid status'})
        
        # Log the change for audit trail
        old_status = session.status
        logger.info(f"Status change for session {session_id}: {old_status} -> {new_status} by {request.user}")
        
        # Update the session
        session.status = new_status
        session.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Status updated from {old_status} to {new_status}'
        })
        
    except GuidanceSession.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Session not found'})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        logger.error(f"Error updating session status: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Server error occurred'})

@login_required
def create_session(request):
    if request.method == 'POST':
        student_number = request.POST.get('studentNumber')
        student_name = request.POST.get('studentName')
        student_email = request.POST.get('studentEmail')
        student_year = request.POST.get('studentYear')
        student_course = request.POST.get('studentCourse')
        student_section = request.POST.get('studentSection')

        offense_type = request.POST.get('offenseType')
        custom_offense = request.POST.get('customOffense')
        offense_details = request.POST.get('offenseDetails')
        incident_date_str = request.POST.get('incidentDate')
        incident_time_str = request.POST.get('incidentTime')
        severity = request.POST.get('severity')
        witnesses = request.POST.get('witnesses')
        reason = request.POST.get('reason', 'disciplinary')
        concern_description = request.POST.get('concernDescription', offense_details)

        # Try to get existing student profile, or create one if it doesn't exist
        try:
            student_profile = Profile.objects.get(user__email=student_email)
        except Profile.DoesNotExist:
            try:
                # Create a new user for the student
                from django.contrib.auth.models import User
                
                # Check if user with this email already exists
                try:
                    existing_user = User.objects.get(email=student_email)
                    # FIXED: If user exists but has no profile, create one for them
                    student_profile = Profile.objects.create(
                        user=existing_user,
                        user_type='student',
                        student_number=student_number,
                        year_level=student_year,
                        course=student_course,
                        section=student_section
                    )
                    messages.info(request, f'Profile created for existing user {student_name}')
                    # Skip the user creation part since we found existing user
                    
                except User.DoesNotExist:
                    # Create new user since none exists with this email
                    # Create new user since none exists with this email
                    
                    # Create new user
                    username = student_email.split('@')[0]  # Use email prefix as username
                    # Make sure username is unique
                    counter = 1
                    original_username = username
                    while User.objects.filter(username=username).exists():
                        username = f"{original_username}{counter}"
                        counter += 1
                    
                    new_user = User.objects.create_user(
                        username=username,
                        email=student_email,
                        first_name=student_name.split()[0] if student_name else '',
                        last_name=' '.join(student_name.split()[1:]) if len(student_name.split()) > 1 else ''
                    )
                    
                    # FIXED: Create profile for the new user with proper fields
                    student_profile = Profile.objects.create(
                        user=new_user,
                        user_type='student',  # Added: Specify user type as student
                        student_number=student_number,  # Added: Set student number
                        year_level=student_year,  # Added: Set year level
                        course=student_course,  # Added: Set course
                        section=student_section  # Added: Set section
                    )
                    
                    messages.info(request, f'New student profile created for {student_name}')
                
            except Exception as e:
                messages.error(request, f'Error creating student profile: {str(e)}')
                return redirect('create_session')

        final_offense_type = custom_offense if offense_type == "other" else offense_type

        # Date/time validation
        try:
            incident_date = datetime.strptime(incident_date_str, '%Y-%m-%d').date() if incident_date_str else None
            incident_time = incident_time_str if incident_time_str else None
        except ValueError as e:
            messages.error(request, f"Invalid date/time format: {e}")
            return redirect('create_session')

        try:
            # 1. Create the GuidanceSession
            guidance_session = GuidanceSession.objects.create(
                student=student_profile,
                student_number=student_number,
                student_name=student_name,
                student_email=student_email,
                student_year=student_year,
                student_course=student_course,
                student_section=student_section,
                offense_type=final_offense_type,
                offense_details=offense_details,
                incident_date=incident_date,
                incident_time=incident_time,
                severity=severity,
                witnesses=witnesses,
                reason=reason,
                concern_description=concern_description,
                status='pending'
            )

            # 2. Create a matching Case (status='pending')
            try:
                faculty_profile = Profile.objects.get(user=request.user)
                Case.objects.create(
                    student=student_profile,
                    counselor=faculty_profile,
                    title=f"Guidance Session: {final_offense_type}",
                    description=concern_description,
                    status='pending'
                )
            except Profile.DoesNotExist:
                messages.warning(request, 'Session created but case could not be assigned - faculty profile not found.')

            messages.success(request, 'Session created successfully!')
            request.session['newSessionCreated'] = True
            return redirect('session_created')

        except Exception as e:
            messages.error(request, f"Error creating session: {str(e)}")
            return redirect('create_session')
    else:
        # GET logic
        context = {
            'user_name': request.user.get_full_name() or request.user.username,
        }
        return render(request, 'app/create_session.html', context)


# Home view
def home(request):
    return render(request, 'app/home.html')

def case_reports(request):
    return render(request, 'app/case_reports.html')

@login_required
def student_cases_dashboard(request):
    try:
        profile = Profile.objects.get(user=request.user)

        # Use select_related to avoid missing data
        all_cases = Case.objects.filter(student=profile).select_related('student__user')
        all_sessions = GuidanceSession.objects.filter(student=profile).select_related('student__user')

        def case_dict(obj, is_session=False):
            try:
                student_name = obj.student.user.get_full_name() or obj.student.user.username
            except:
                student_name = "Unknown Student"

            if is_session:
                return {
                    'id': obj.id,
                    'type': 'session',
                    'title': f"Guidance Session - {getattr(obj, 'get_reason_display', lambda: obj.reason)()}",
                    'status': 'in_progress' if obj.status == 'approved' else (
                        'resolved' if obj.status == 'completed' else obj.status),
                    'student_name': student_name,
                    'description': obj.concern_description,
                }
            else:
                return {
                    'id': obj.id,
                    'type': 'case',
                    'title': obj.title,
                    'status': obj.status,
                    'student_name': student_name,
                    'description': obj.description,
                }

        # Combine into labeled categories
        pending_cases = [case_dict(c) for c in all_cases.filter(status='pending')] \
                        + [case_dict(s, True) for s in all_sessions.filter(status='pending')]

        ongoing_cases = [case_dict(c) for c in all_cases.filter(status__in=['in_progress', 'under_review'])] \
                        + [case_dict(s, True) for s in all_sessions.filter(status='approved')]

        completed_cases = [case_dict(c) for c in all_cases.filter(status='resolved')] \
                        + [case_dict(s, True) for s in all_sessions.filter(status='completed')]

        case_history = [case_dict(c) for c in all_cases] + [case_dict(s, True) for s in all_sessions]

        context = {
            'pending_cases': pending_cases,
            'ongoing_cases': ongoing_cases,
            'completed_cases': completed_cases,
            'case_history': case_history,
            'user_name': request.user.first_name or request.user.username,
        }

        return render(request, 'app/student_cases_dashboard.html', context)

    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')





def case_history(request):
    return render(request, 'app/case_history.html')

def case_detail(request):
    return render(request, 'app/case_detail.html')

@login_required
def session_created(request):
    # You can filter by faculty if needed (e.g., assigned_counselor=request.user.profile)
    all_sessions = GuidanceSession.objects.all().order_by('-created_at')

    # Separate by status for the tabs
    ongoing_sessions = all_sessions.filter(status='approved')
    pending_sessions = all_sessions.filter(status='pending')
    completed_sessions = all_sessions.filter(status='completed')
    history_sessions = all_sessions.filter()

    context = {
        'user_name': request.user.get_full_name() or request.user.username,
        'ongoing_sessions': ongoing_sessions,
        'pending_sessions': pending_sessions,
        'completed_sessions': completed_sessions,
        'history_sessions': history_sessions,
    }
    return render(request, 'app/session_created.html', context)






@login_required
def dashboard(request):
    """Legacy dashboard view: redirect to the new dashboards and avoid old templates."""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type == 'counselor':
            return redirect('counselor_dashboard')
        else:
            return redirect('student_dashboard')
    except Profile.DoesNotExist:
        return redirect('home')


# About view
def about(request):
    return render(request, 'app/about.html')

def contact(request):
    return render(request, 'app/contact.html')



def verify_otp(request, user_id):
    """OTP verification view for non-CVSU email signups"""
    try:
        user = User.objects.get(id=user_id)
        verification = EmailVerification.objects.get(user=user)
        
        if request.method == 'POST':
            otp_input = request.POST.get('otp', '').strip()
            
            # Check if OTP is expired
            if verification.is_expired():
                messages.error(request, "OTP has expired. Please request a new one.")
                return render(request, 'app/verify_otp.html', {
                    'user_id': user_id,
                    'email': verification.email,
                    'expired': True
                })
            
            # Verify OTP
            if verification.otp == otp_input:
                # Mark as verified and activate user
                verification.verified = True
                verification.save()
                
                user.is_active = True
                user.save()
                
                messages.success(request, "Email successfully verified! You can now log in.")
                return redirect('login')
            else:
                messages.error(request, "Invalid OTP. Please try again.")
        
        context = {
            'user_id': user_id,
            'email': verification.email,
            'expired': False
        }
        return render(request, 'app/verify_otp.html', context)
    
    except User.DoesNotExist:
        messages.error(request, "User not found.")
        return redirect('signup')
    except EmailVerification.DoesNotExist:
        messages.error(request, "Verification record not found.")
        return redirect('signup')
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('signup')


def resend_otp(request, user_id):
    """Resend OTP for email verification"""
    try:
        user = User.objects.get(id=user_id)
        verification = EmailVerification.objects.get(user=user)
        
        from .otp_utils import generate_otp, send_otp_email
        
        # Generate new OTP
        otp = generate_otp()
        verification.otp = otp
        verification.created_at = timezone.now()
        verification.save()
        
        # Send OTP email
        success, message = send_otp_email(verification.email, otp, user.username)
        
        if success:
            messages.success(request, 'New OTP sent to your email!')
        else:
            messages.error(request, f'Failed to send OTP: {message}')
        
        return redirect('verify_otp', user_id=user_id)
    
    except (User.DoesNotExist, EmailVerification.DoesNotExist):
        messages.error(request, "Verification record not found.")
        return redirect('signup')




# Sign up view - Updated with OTP for non-CVSU emails
def signup(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        user_type = request.POST.get('user_type', 'student')
        student_number = request.POST.get('student_number', '').strip()
        counselor_id = request.POST.get('counselor_id', '').strip()

        if user_type not in ['student', 'counselor']:
            messages.error(request, 'Invalid account type selected.')
            return redirect('signup')

        # Validate password match
        if password1 != password2:
            messages.error(request, "Passwords do not match.")
            return redirect('signup')

        # Check if username already exists
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return redirect('signup')

        # Check if email already exists
        if User.objects.filter(email=email).exists():
            messages.error(request, 'An account with this email already exists.')
            return redirect('signup')

        # Validate role-specific ID fields
        if user_type == 'student':
            if not student_number:
                messages.error(request, "Student number is required for student accounts.")
                return redirect('signup')
            if Profile.objects.filter(student_number=student_number).exists():
                messages.error(request, "This student number is already registered.")
                return redirect('signup')
        else:
            # Testing mode: counselor_id is optional and not uniqueness-validated.
            pass

        # Validate email
        try:
            validate_email(email)
        except ValidationError:
            messages.error(request, "Invalid email address.")
            return redirect('signup')

        # Check if CVSU email - auto-verify
        is_cvsu_email = email.endswith('@cvsu.edu.ph')
        
        # Create user (inactive until verified for non-CVSU emails)
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password1,
            is_active=is_cvsu_email  # CVSU emails are auto-active
        )

        # Create profile based on selected account type
        profile = Profile.objects.create(
            user=user,
            user_type=user_type,
            student_number=student_number if user_type == 'student' else None,
            counselor_id=counselor_id if user_type == 'counselor' else None,
        )

        if is_cvsu_email:
            # CVSU email - no OTP needed
            messages.success(request, f'Account created for {username}! You can now log in.')
            return redirect('login')
        else:
            # Non-CVSU email - send OTP
            from .otp_utils import generate_otp, send_otp_email
            
            otp = generate_otp()
            
            # Create or update email verification record
            EmailVerification.objects.update_or_create(
                user=user,
                defaults={
                    'otp': otp,
                    'email': email,
                    'verified': False
                }
            )
            
            # Send OTP email
            success, message = send_otp_email(email, otp, username)
            
            if success:
                messages.success(request, 'Account created! Please check your email for the OTP to verify your account.')
                return redirect('verify_otp', user_id=user.id)
            else:
                messages.error(request, f'Account created but failed to send OTP: {message}. Please contact support.')
                return redirect('signup')

    return render(request, 'app/signup.html')




def login_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        User = get_user_model()
        try:
            user_obj = User.objects.get(email=email)
            authenticated_user = authenticate(request, username=user_obj.username, password=password)
            
            if authenticated_user is not None:
                login(request, authenticated_user)
                
                # Kunin ang profile at alamin ang user_type
                profile = Profile.objects.get(user=authenticated_user)
                
                # Redirect ayon sa user_type
                if profile.user_type == 'counselor':
                    return redirect('counselor_dashboard')  # Counselor → New Counselor Dashboard
                else:
                    return redirect('student_dashboard')  # Student → Student Dashboard
            else:
                messages.error(request, 'Invalid password')
        except User.DoesNotExist:
            messages.error(request, 'No account found with this email')
        except Profile.DoesNotExist:
            messages.error(request, 'Profile not found. Please contact support.')
    
    return render(request, 'app/login.html')

@login_required
def student_list(request):
    # Only allow faculty or counselors to access this page
    profile = Profile.objects.get(user=request.user)
    if profile.user_type != 'counselor':  # or adjust this depending on your app
        messages.error(request, "You do not have access to view this page.")
        return redirect('student_dashboard')

    # Get all students
    students = Profile.objects.filter(user_type='student').prefetch_related(
        'user',
        'sessions',  # reverse relationship from Profile to GuidanceSession
        'cases'  # reverse relationship from Profile to Case
    )

    context = {
        'students': students,
        'user_name': request.user.get_full_name() or request.user.username,
        'case_ids': [student.case.id for student in students]
    }
    return render(request, 'app/student_list.html', context)


# Student dashboard view - UPDATED to include guidance sessions as cases
@login_required
def student_dashboard(request):
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'student':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('dashboard')  # ← Faculty? Redirect sa faculty dashboard
        
        user = request.user
        
        # Fetch the student's regular cases
        regular_cases = Case.objects.filter(student=profile).order_by('-created_at')
        
        # Fetch the student's guidance sessions and convert them to case-like objects
        guidance_sessions = GuidanceSession.objects.filter(student=profile).order_by('-created_at')
        
        # Combine cases and guidance sessions
        all_cases = []
        
        # Add regular cases
        for case in regular_cases:
            all_cases.append({
                'id': case.id,
                'title': case.title,
                'type': 'case',
                'status': case.status,
                'created_at': case.created_at,
                'counselor': case.counselor,
                'description': case.description
            })
        
        # Add guidance sessions as cases
        for session in guidance_sessions:
            # Map guidance session status to case status
            case_status = 'pending' if session.status == 'pending' else session.status
            if session.status == 'approved':
                case_status = 'in_progress'
            elif session.status == 'completed':
                case_status = 'resolved'
            
            all_cases.append({
                'id': session.id,
                'title': f"Guidance Session - {session.get_reason_display()}",
                'type': 'guidance_session',
                'status': case_status,
                'created_at': session.created_at,
                'counselor': session.assigned_counselor or session.preferred_counselor,
                'description': session.concern_description,
                'session_status': session.status
            })
        
        # Sort all cases by creation date (most recent first)
        all_cases.sort(key=lambda x: x['created_at'], reverse=True)
        
        # Get latest 3 cases for dashboard
        recent_cases = all_cases[:3]
        
        # Get counts for statistics (combining both types)
        regular_active = Case.objects.filter(student=profile, status__in=['in_progress', 'under_review']).count()
        session_active = GuidanceSession.objects.filter(student=profile, status__in=['approved']).count()
        active_cases = regular_active + session_active
        
        regular_resolved = Case.objects.filter(student=profile, status='resolved').count()
        session_resolved = GuidanceSession.objects.filter(student=profile, status='completed').count()
        resolved_cases = regular_resolved + session_resolved
        
        regular_pending = Case.objects.filter(student=profile, status='pending').count()
        session_pending = GuidanceSession.objects.filter(student=profile, status='pending').count()
        pending_cases = regular_pending + session_pending
    # Get notifications
        notif_data = get_notifications(request)
        context = {
            'profile': profile,
            'user_name': user.first_name if user.first_name else user.username,
            'user_email': user.email,
            'all_cases': recent_cases,
            'active_cases': active_cases,
            'resolved_cases': resolved_cases,
            'pending_cases': pending_cases,
            'total_cases': len(all_cases),
            'guidance_sessions': guidance_sessions,
            'notifications': notif_data['notifications'],
            'unread_count': notif_data['unread_count'],
        }
        return render(request, 'app/student_dashboard.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')

        

# Student cases view - UPDATED to show both cases and guidance sessions
@login_required
def student_cases(request):
    try:
        profile = Profile.objects.get(user=request.user)
        
        # Fetch the student's regular cases
        regular_cases = Case.objects.filter(student=profile).order_by('-created_at')
        
        # Fetch the student's guidance sessions
        guidance_sessions = GuidanceSession.objects.filter(student=profile).order_by('-created_at')
        
        # Combine cases and guidance sessions
        all_cases = []
        
        # Add regular cases
        for case in regular_cases:
            all_cases.append({
                'id': case.id,
                'title': case.title,
                'type': 'case',
                'status': case.status,
                'created_at': case.created_at,
                'counselor': case.counselor,
                'description': case.description
            })
        
        # Add guidance sessions as cases
        for session in guidance_sessions:
            # Map guidance session status to case status
            case_status = 'pending' if session.status == 'pending' else session.status
            if session.status == 'approved':
                case_status = 'in_progress'
            elif session.status == 'completed':
                case_status = 'resolved'
            
            all_cases.append({
                'id': session.id,
                'title': f"Guidance Session - {session.get_reason_display()}",
                'type': 'guidance_session',
                'status': case_status,
                'created_at': session.created_at,
                'counselor': session.assigned_counselor or session.preferred_counselor,
                'description': session.concern_description,

                'session_status': session.status
            })
        
        # Sort all cases by creation date (most recent first)
        all_cases.sort(key=lambda x: x['created_at'], reverse=True)
        
        context = {
            'cases': all_cases,
            'user_name': request.user.first_name if request.user.first_name else request.user.username
        }
        return render(request, 'app/student_cases.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')

# Create comment view
@login_required
def create_comment(request, case_id):
    try:
        profile = Profile.objects.get(user=request.user)
        case = get_object_or_404(Case, id=case_id)
        
        # Check if the user is related to this case (either as student or counselor)
        if case.student != profile and case.counselor != profile:
            messages.error(request, "You don't have permission to comment on this case.")
            return redirect('student_dashboard')
        
        if request.method == 'POST':
            form = CommentForm(request.POST)
            if form.is_valid():
                comment = form.save(commit=False)
                comment.case = case
                comment.user = request.user
                comment.save()
                
                messages.success(request, "Your comment has been added successfully.")
                return redirect('view_case_details', case_id=case.id)
        else:
            form = CommentForm()
        
        return render(request, 'app/create_comment.html', {
            'form': form,
            'case': case
        })
    except (Profile.DoesNotExist, Case.DoesNotExist):
        messages.error(request, "Case not found or you don't have permission to comment.")
        return redirect('student_dashboard')


@login_required
def schedule_guidance_session(request):
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'student':
            messages.error(request, "Only students can schedule guidance sessions")
            return redirect('home')
    except Profile.DoesNotExist:
        messages.error(request, "Profile not found")
        return redirect('home')

    # Check if this session is for a specific case
    case_id = request.GET.get('case_id')
    linked_case = None
    if case_id:
        try:
            linked_case = Case.objects.get(id=case_id, student=profile)
        except Case.DoesNotExist:
            messages.error(request, "Case not found.")
            return redirect('student_cases_only')

    counselors = Profile.objects.filter(user_type='counselor')
    
    if request.method == 'POST':
        form_data = request.POST.copy()
        form = GuidanceSessionForm(form_data)
        
        if form.is_valid():
            from app.utils import NotificationManager
            
            session = form.save(commit=False)
            session.student = profile
            session.status = 'pending'
            session.save()

            # Notify counselor (or preferred counselor)
            counselor_to_notify = session.preferred_counselor or session.assigned_counselor
            if counselor_to_notify:
                NotificationManager.notify_session_requested(session, counselor_to_notify)
            else:
                # Notify all counselors if no specific counselor selected
                all_counselors = Profile.objects.filter(user_type='counselor')
                for counselor in all_counselors:
                    NotificationManager.notify_session_requested(session, counselor)

            # If this session is for a specific case, link it
            if linked_case:
                linked_case.sessions.add(session)
                messages.success(request, f"Session request submitted and linked to Case #{linked_case.id}! You can view it in 'My Cases'.")
                return redirect('student_cases_only')
            else:
                # Standalone session - counselor will decide if it needs a formal case
                messages.success(request, "Your guidance session request has been submitted successfully! You can view it in 'My Sessions'.")
                return redirect('student_sessions_only')
            
            request.session['guidance_session_id'] = session.id
        else:
            print(form.errors)
    else:
        # Pre-fill form if linked to a case
        initial_data = {}
        if linked_case:
            initial_data = {
                'concern_description': f"Follow-up for Case #{linked_case.id}: {linked_case.title}\n\n{linked_case.description}\n\n---\nAdditional concerns:",
                'preferred_counselor': linked_case.counselor.id if linked_case.counselor else None,
            }
        form = GuidanceSessionForm(initial=initial_data)
    
    return render(request, 'app/schedule_guidance_session.html', {
        'form': form,
        'counselors': counselors,
        'linked_case': linked_case,
        'user_name': request.user.first_name if request.user.first_name else request.user.username
    })

@login_required
def student_case_view(request):
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'student':
            messages.error(request, "Only students can view this page.")
            return redirect('home')
    except Profile.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('home')

    # Get all cases for the logged-in student with proper ordering
    all_cases = Case.objects.filter(student=profile).order_by('-created_at')
    all_sessions = GuidanceSession.objects.filter(student=profile).order_by('-created_at')

    # Initialize empty lists
    pending_cases = []
    ongoing_cases = []
    completed_cases = []
    case_history = []

    # Process Cases
    for case in all_cases:
        case_data = case_dict(case, is_session=False)
        case_history.append(case_data)
        
        if case.status == 'pending':
            pending_cases.append(case_data)
        elif case.status in ['in_progress', 'under_review']:
            ongoing_cases.append(case_data)
        elif case.status == 'resolved':
            completed_cases.append(case_data)

    # Process Guidance Sessions
    for session in all_sessions:
        session_data = case_dict(session, is_session=True)
        case_history.append(session_data)
        
        if session.status == 'pending':
            pending_cases.append(session_data)
        elif session.status == 'approved':
            ongoing_cases.append(session_data)
        elif session.status == 'completed':
            completed_cases.append(session_data)

    # Sort all lists by created_at (newest first)
    pending_cases.sort(key=lambda x: x['created_at'], reverse=True)
    ongoing_cases.sort(key=lambda x: x['created_at'], reverse=True)
    completed_cases.sort(key=lambda x: x['created_at'], reverse=True)
    case_history.sort(key=lambda x: x['created_at'], reverse=True)

    # Debug information (remove in production)
    print(f"Debug - Total cases: {len(all_cases)}")
    print(f"Debug - Total sessions: {len(all_sessions)}")
    print(f"Debug - Pending: {len(pending_cases)}")
    print(f"Debug - Ongoing: {len(ongoing_cases)}")
    print(f"Debug - Completed: {len(completed_cases)}")

    return render(request, 'app/student_case_dashboard.html', {
        'user_name': request.user.first_name or request.user.username,
        'pending_cases': pending_cases,
        'ongoing_cases': ongoing_cases,
        'completed_cases': completed_cases,
        'case_history': case_history,
        'total_cases': len(all_cases),
        'total_sessions': len(all_sessions),
    })

# Helper function to convert cases and sessions into dictionaries for rendering
def case_dict(instance, is_session=False):
    """
    Convert Case or GuidanceSession instance to dictionary for template rendering
    """
    try:
        if is_session:
            return {
                'id': instance.id,
                'title': getattr(instance, 'reason', 'No reason provided'),
                'status': getattr(instance, 'status', 'unknown'),
                'created_at': getattr(instance, 'created_at', timezone.now()),
                'case_type': 'guidance_session',
                'description': getattr(instance, 'description', ''),  # Additional field
            }
        else:
            return {
                'id': instance.id,
                'title': getattr(instance, 'title', 'Untitled Case'),
                'status': getattr(instance, 'status', 'unknown'),
                'created_at': getattr(instance, 'created_at', timezone.now()),
                'case_type': 'case',
                'description': getattr(instance, 'description', ''),  # Additional field
            }
    except Exception as e:
        # Log the error and return a safe default
        print(f"Error in case_dict: {e}")
        return {
            'id': getattr(instance, 'id', 0),
            'title': 'Error loading case',
            'status': 'error',
            'created_at': timezone.now(),
            'case_type': 'guidance_session' if is_session else 'case',
            'description': '',
        }








@csrf_exempt
@login_required
def update_session_time(request, session_id):
    if request.method == 'POST':
        try:
            # Fetch the session object
            session = GuidanceSession.objects.get(id=session_id)

            # Check if the session is pending
            if session.status != 'pending':
                return JsonResponse({'success': False, 'error': 'Only pending sessions can have time set'})

            # Get the new session time from the request
            new_session_time = request.POST.get('session_time')

            # Update the session time (set the preferred date/time for pending session)
            session.preferred_date = new_session_time
            session.save()

            return JsonResponse({'success': True})
        except GuidanceSession.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Session not found'})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})





@login_required
def set_session_schedule(request, session_id):
    if request.user.profile.user_type != 'counselor':
        messages.error(request, "Only counselors can set session schedules.")
        return redirect('dashboard')

    session = get_object_or_404(GuidanceSession, id=session_id)

    if session.status != 'pending':
        messages.error(request, "This session cannot be scheduled.")
        return redirect('dashboard')

   # if request.method == 'POST':
       # form = SetScheduleForm(request.POST, instance=session)
        if form.is_valid():
            session = form.save(commit=False)
            session.assigned_counselor = request.user.profile
            session.status = 'approved'
            session.save()
            messages.success(request, "Session schedule has been set!")
            return redirect('dashboard')
   # else:
       # form = SetScheduleForm(instance=session)

  #  return render(request, 'app/set_session_schedule.html', {'form': form, 'session': session})



@login_required
def guidance_request_success(request):
    # Get the session ID from session variable
    session_id = request.session.get('guidance_session_id')
    
    try:
        # Get the session object
        session = GuidanceSession.objects.get(id=session_id, student__user=request.user)
        
        # Clear the session variable
        if 'guidance_session_id' in request.session:
            del request.session['guidance_session_id']
        
        return render(request, 'app/guidance_request_success.html', {
            'session': session
        })
    except GuidanceSession.DoesNotExist:
        # If the session doesn't exist or doesn't belong to the user, redirect to dashboard
        messages.error(request, "Session information not found.")
        return redirect('student_dashboard')


@login_required
def check_case_status(request):
    # Get user's first name for sidebar
    user_name = request.user.first_name or request.user.username

    # Get notifications
    notif_data = get_notifications(request)

    context = {
        'form_submitted': False,
        'case_found': False,
        'case': None,
        'case_type': None,
        'user_name': user_name,
        'notifications': notif_data['notifications'],
        'unread_count': notif_data['unread_count'],
    }


    if request.method == 'POST':
        case_number = request.POST.get('case_number', '').strip()
        context['form_submitted'] = True

        try:
            # Get the user's profile
            profile = Profile.objects.get(user=request.user)
            
            # First, try to find a regular Case with this ID
            try:
                case = Case.objects.get(id=case_number, student=profile)
                context['case'] = case
                context['case_found'] = True
                context['case_type'] = 'regular_case'
                
                # Get comments/updates for this case
                context['comments'] = Comment.objects.filter(case=case).order_by('-created_at')
                
            except Case.DoesNotExist:
                # If not found in Case, try GuidanceSession
                try:
                    guidance_session = GuidanceSession.objects.get(id=case_number, student=profile)
                    context['case'] = guidance_session
                    context['case_found'] = True
                    context['case_type'] = 'guidance_session'
                    
                except GuidanceSession.DoesNotExist:
                    # If not found in either model
                    messages.error(request, "No case found with that number for your account.")
                    
        except Profile.DoesNotExist:
            messages.error(request, "Profile not found.")

    return render(request, 'app/check_case_status.html', context)

@login_required
def view_case_details(request, case_id):
    try:
        # Ensure the case belongs to the current student
        profile = Profile.objects.get(user=request.user)
        case = Case.objects.get(id=case_id, student=profile)
        
        # Get comments for this case
        comments = Comment.objects.filter(case=case).order_by('-created_at')
        
        context = {
            'case': case,
            'comments': comments
        }
        return render(request, 'app/case_details.html', context)
    except (Profile.DoesNotExist, Case.DoesNotExist):
        messages.error(request, "Case not found or you don't have permission to view it.")
        return redirect('student_dashboard')
    
    

# NEW VIEW: View guidance session details
@login_required
def view_guidance_session_details(request, session_id):
    try:
        # Ensure the session belongs to the current student
        profile = Profile.objects.get(user=request.user)
        session = GuidanceSession.objects.get(id=session_id, student=profile)
        
        context = {
            'session': session
        }
        return render(request, 'app/guidance_session_details.html', context)
    except (Profile.DoesNotExist, GuidanceSession.DoesNotExist):
        messages.error(request, "Guidance session not found or you don't have permission to view it.")
        return redirect('student_dashboard')





import json
import logging
from django.http import JsonResponse, HttpResponse, FileResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Notification, Case, Profile

# Setup Logger
logger = logging.getLogger(__name__)


logger = logging.getLogger(__name__)

@login_required
@require_http_methods(["POST"])
def update_case_status(request, case_id):
    """
    Enhanced view to update case status with better error handling
    """
    try:
        logger.info(f"=== UPDATE CASE STATUS REQUEST ===")
        logger.info(f"Case ID: {case_id}")
        logger.info(f"User: {request.user}")
        logger.info(f"Method: {request.method}")
        logger.info(f"Content Type: {request.content_type}")
        
        # Validate case_id format
        try:
            case_id_int = int(case_id)
            if case_id_int <= 0:
                raise ValueError("Case ID must be positive")
        except (ValueError, TypeError) as e:
            logger.error(f"Invalid case_id format: {case_id} - {str(e)}")
            return JsonResponse({
                'success': False, 
                'error': f'Invalid case ID format: {case_id}'
            }, status=400)
        
        # Parse request body
        try:
            if request.content_type == 'application/json':
                data = json.loads(request.body)
            else:
                data = request.POST
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON in request body: {str(e)}")
            return JsonResponse({
                'success': False, 
                'error': 'Invalid JSON data'
            }, status=400)
        
        logger.info(f"Request data: {data}")
        
        # Get the case with detailed error handling
        try:
            # First, let's see if any cases exist
            total_cases = Case.objects.count()
            logger.info(f"Total cases in database: {total_cases}")
            
            # Try to get the specific case
            case = Case.objects.get(id=case_id)
            logger.info(f"Found case: {case}")
            
        except Case.DoesNotExist:
            # Enhanced debugging for missing case
            logger.error(f"=== CASE NOT FOUND DEBUG ===")
            logger.error(f"Looking for case_id: {case_id_int}")
            
            # Show what cases actually exist
            existing_cases = Case.objects.values_list('id', flat=True)[:10]
            logger.error(f"Existing case IDs (first 10): {list(existing_cases)}")
            
            # Check if it's a soft-deleted case or has different conditions
            all_cases_query = Case.objects.all()
            logger.error(f"Total cases query count: {all_cases_query.count()}")
            
            # If using soft deletes, check deleted cases too
            try:
                # Adjust this based on your model structure
                if hasattr(Case, 'is_deleted'):
                    deleted_case = Case.objects.filter(id=case_id_int, is_deleted=True).first()
                    if deleted_case:
                        logger.error(f"Case {case_id} exists but is soft-deleted")
                        return JsonResponse({
                            'success': False,
                            'error': f'Case {case_id} has been deleted'
                        }, status=404)
            except Exception as e:
                logger.error(f"Error checking soft deletes: {str(e)}")
            
            return JsonResponse({
                'success': False,
                'error': f'Case with ID {case_id} not found'
            }, status=404)
        
        except Exception as e:
            logger.error(f"Unexpected error getting case: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'Database error: {str(e)}'
            }, status=500)
        
        # Validate user permissions (adjust based on your auth logic)
        # Example: Check if user owns this case or has permission
        try:
            # Add your permission checks here
            # if case.student.user != request.user and not request.user.is_staff:
            #     return JsonResponse({
            #         'success': False,
            #         'error': 'Permission denied'
            #     }, status=403)
            pass
        except Exception as e:
            logger.error(f"Permission check error: {str(e)}")
        
        # Extract and validate data
        new_status = data.get('status', '').strip()
        scheduled_date = data.get('scheduled_date', '').strip()
        scheduled_time = data.get('scheduled_time', '').strip()
        
        logger.info(f"Updating case {case_id_int}:")
        logger.info(f"  New status: '{new_status}'")
        logger.info(f"  Scheduled date: '{scheduled_date}'")
        logger.info(f"  Scheduled time: '{scheduled_time}'")
        
        # Validate status
        VALID_STATUSES = ['pending', 'in_progress', 'completed', 'cancelled']  # Adjust based on your model
        if new_status not in VALID_STATUSES:
            return JsonResponse({
                'success': False,
                'error': f'Invalid status. Must be one of: {", ".join(VALID_STATUSES)}'
            }, status=400)
        
        # Store old values for logging
        old_status = case.status
        old_scheduled_date = getattr(case, 'scheduled_date', None)
        old_scheduled_time = getattr(case, 'scheduled_time', None)
        
        # Update the case
        try:
            case.status = new_status
            
            # Handle scheduled date/time if your model has these fields
            if hasattr(case, 'scheduled_date'):
                case.scheduled_date = scheduled_date if scheduled_date else None
            if hasattr(case, 'scheduled_time'):
                case.scheduled_time = scheduled_time if scheduled_time else None
            
            # Set updated timestamp if you have one
            if hasattr(case, 'updated_at'):
                from django.utils import timezone
                case.updated_at = timezone.now()
            
            case.save()
            
            logger.info(f"✅ Case {case_id_int} updated successfully")
            logger.info(f"  Status: {old_status} → {new_status}")
            logger.info(f"  Date: {old_scheduled_date} → {scheduled_date}")
            logger.info(f"  Time: {old_scheduled_time} → {scheduled_time}")
            
        except Exception as e:
            logger.error(f"Error saving case: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'Failed to save case: {str(e)}'
            }, status=500)
        
        # Return success response
        response_data = {
            'success': True,
            'message': 'Case updated successfully',
            'case_id': case_id_int,
            'new_status': new_status
        }
        
        # Include updated fields in response if needed
        if scheduled_date:
            response_data['scheduled_date'] = scheduled_date
        if scheduled_time:
            response_data['scheduled_time'] = scheduled_time
        
        logger.info(f"Returning success response: {response_data}")
        return JsonResponse(response_data)
        
    except Exception as e:
        logger.error(f"=== UNEXPECTED ERROR ===")
        logger.error(f"Error type: {type(e).__name__}")
        logger.error(f"Error message: {str(e)}")
        logger.exception("Full traceback:")
        
        return JsonResponse({
            'success': False,
            'error': f'Server error: {str(e)}'
        }, status=500)


# Optional: Add a debug view to check what cases exist
@login_required
def debug_cases(request):
    """
    Debug view to see what cases exist - REMOVE IN PRODUCTION
    """
    if not request.user.is_staff:  # Only allow admin users
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        cases = Case.objects.all().values('id', 'status', 'created_at')[:20]
        return JsonResponse({
            'total_cases': Case.objects.count(),
            'sample_cases': list(cases)
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)




@login_required
def student_case(request):

    try:
        profile = Profile.objects.get(user=request.user)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found.")
        return redirect('home')
    

    if profile.user_type != 'counselor':
        messages.error(request, "Only faculty members can access this page.")
        return redirect('dashboard')

    all_cases = Case.objects.select_related('student__user').all()
    all_sessions = GuidanceSession.objects.select_related('student__user').all()

    def case_dict(obj, is_session=False):
        # Build a dict with all fields template needs
        if is_session:
            return {
                'title': f"Guidance Session - {obj.get_reason_display() if hasattr(obj, 'get_reason_display') else obj.reason}",
                'status': 'in_progress' if obj.status == 'approved' else ('resolved' if obj.status == 'completed' else obj.status),
                'student_name': obj.student.user.get_full_name() or obj.student.user.username,
                'description': obj.concern_description,
            }
        else:
            return {
                'title': obj.title,
                'status': obj.status,
                'student_name': obj.student.user.get_full_name() or obj.student.user.username,
                'description': obj.description,
            }

    # Now merge both regular cases and sessions into each bucket
   
    
    pending_cases = [case_dict(c) for c in all_cases.filter(status='pending')] \
                 + [case_dict(s, True) for s in all_sessions.filter(status='pending')]
    ongoing_cases = [case_dict(c) for c in all_cases.filter(status__in=['in_progress', 'under_review'])] \
                 + [case_dict(s, True) for s in all_sessions.filter(status='approved')]
    completed_cases = [case_dict(c) for c in all_cases.filter(status='resolved')] \
                 + [case_dict(s, True) for s in all_sessions.filter(status='completed')]
    case_history = [case_dict(c) for c in all_cases] + [case_dict(s, True) for s in all_sessions]

    return render(request, 'app/student_case.html',  {
        'user_name': request.user.first_name or request.user.username,
        'pending_cases': pending_cases,
        'ongoing_cases': ongoing_cases,
        'completed_cases': completed_cases,
        'case_history': case_history,
    })






@login_required
def student_list(request):
    try:
        # Check if the user is faculty
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('student_dashboard')
        
        # Get all student profiles with their related data
        students = Profile.objects.filter(user_type='student').select_related('user').prefetch_related(
            'cases',  # Regular cases
            'sessions'  # Guidance sessions
        ).order_by('user__first_name', 'user__last_name', 'user__username')
        
        user = request.user
        context = {
            'students': students,
            'user_name': request.user.get_full_name() or request.user.username,
            
        }
        return render(request, 'app/student_list.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')




@login_required
def analytics_dashboard(request):
    # Calculate actual statistics from your models
    total_cases = Case.objects.count() + GuidanceSession.objects.count()
    active_count = Case.objects.filter(status__in=['in_progress', 'under_review']).count() + GuidanceSession.objects.filter(status='approved').count()
    completed_count = Case.objects.filter(status='resolved').count() + GuidanceSession.objects.filter(status='completed').count()
    
    context = {
        'total_cases': total_cases,
        'active_count': active_count,
        'completed_count': completed_count,
    }
    return render(request, 'app/analytics.html', context)


@login_required
def evaluation(request):
    """Main evaluation page - displays form and recent evaluations"""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('student_dashboard')

        # Get recent evaluations (non-draft only)
        recent_evaluations = StudentEvaluation.objects.filter(
            is_draft=False
        ).select_related('student__user', 'evaluator__user').order_by('-created_at')[:5]

        # ✅ Only students who have submitted a case or requested session
        from django.db.models import Q
        students = Profile.objects.filter(
            user_type='student'
        ).filter(
            Q(cases__isnull=False) | Q(sessions__isnull=False)
        ).distinct().order_by('user__first_name', 'user__last_name')

        # ✅ Pass to template
        context = {
            'recent_evaluations': recent_evaluations,
            'students': students,
            'user_name': request.user.first_name if request.user.first_name else request.user.username,
        }

        return render(request, 'app/evaluation.html', context)

    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')
    

@login_required
def submit_evaluation(request):
    """Handle evaluation form submission via AJAX"""
    if request.method == 'POST':
        try:
            profile = Profile.objects.get(user=request.user)
            if profile.user_type != 'counselor':
                return JsonResponse({'success': False, 'error': 'Permission denied'})
            
            # Get form data
            student_id = request.POST.get('student')
            evaluation_date = request.POST.get('evaluation_date')
            reason = request.POST.get('reason')
            hearing = request.POST.get('hearing')
            assessment = request.POST.get('assessment')
            severity = request.POST.get('severity')
            follow_up = request.POST.get('follow_up', 'none')
            
            # Validate required fields
            if not all([student_id, evaluation_date, reason, hearing, assessment, severity]):
                return JsonResponse({'success': False, 'error': 'All required fields must be filled'})
            
            # Get student profile
            student = Profile.objects.get(id=student_id, user_type='student')
            
            # Create evaluation
            evaluation = StudentEvaluation.objects.create(
                student=student,
                evaluator=profile,
                evaluation_date=evaluation_date,
                reason_for_session=reason,
                hearing_frequency=hearing,
                detailed_assessment=assessment,
                severity_level=severity,
                follow_up_required=follow_up,
                is_draft=False
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Evaluation submitted successfully!',
                'evaluation_id': evaluation.id
            })
            
        except Profile.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Student not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def save_evaluation_draft(request):
    """Save evaluation as draft"""
    if request.method == 'POST':
        try:
            profile = Profile.objects.get(user=request.user)
            if profile.user_type != 'counselor':
                return JsonResponse({'success': False, 'error': 'Permission denied'})
            
            # Get form data
            student_id = request.POST.get('student')
            evaluation_date = request.POST.get('evaluation_date')
            reason = request.POST.get('reason')
            hearing = request.POST.get('hearing')
            assessment = request.POST.get('assessment')
            severity = request.POST.get('severity')
            follow_up = request.POST.get('follow_up', 'none')
            
            # Basic validation (only student is required for draft)
            if not student_id:
                return JsonResponse({'success': False, 'error': 'Student must be selected'})
            
            student = Profile.objects.get(id=student_id, user_type='student')
            
            # Create or update draft
            evaluation = StudentEvaluation.objects.create(
                student=student,
                evaluator=profile,
                evaluation_date=evaluation_date or timezone.now().date(),
                reason_for_session=reason or 'other',
                hearing_frequency=hearing or '1st',
                detailed_assessment=assessment or '',
                severity_level=severity or 'minor',
                follow_up_required=follow_up,
                is_draft=True
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Draft saved successfully!',
                'draft_id': evaluation.id
            })
            
        except Profile.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Student not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def get_recent_evaluations(request):
    """Get recent evaluations for the sidebar"""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            return JsonResponse({'success': False, 'error': 'Permission denied'})
        
        evaluations = StudentEvaluation.objects.filter(
            is_draft=False
        ).select_related('student__user').order_by('-created_at')[:5]
        
        evaluation_list = []
        for eval in evaluations:
            evaluation_list.append({
                'id': eval.id,
                'student_name': eval.student.user.get_full_name() or eval.student.user.username,
                'date': eval.evaluation_date.strftime('%B %d, %Y'),
                'reason': eval.get_reason_for_session_display(),
                'hearing': eval.get_hearing_frequency_display(),
                'severity': eval.get_severity_level_display().title()
            })
        
        return JsonResponse({
            'success': True,
            'evaluations': evaluation_list
        })
        
    except Profile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Profile not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def evaluation_list(request):
    """Display all evaluations with filtering options"""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('student_dashboard')
        
        # Get all evaluations
        evaluations = StudentEvaluation.objects.filter(
            is_draft=False
        ).select_related('student__user', 'evaluator__user').order_by('-created_at')
        
        # Apply filters if provided
        student_filter = request.GET.get('student')
        severity_filter = request.GET.get('severity')
        hearing_filter = request.GET.get('hearing')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        if student_filter:
            evaluations = evaluations.filter(student__id=student_filter)
        if severity_filter:
            evaluations = evaluations.filter(severity_level=severity_filter)
        if hearing_filter:
            evaluations = evaluations.filter(hearing_frequency=hearing_filter)
        if date_from:
            evaluations = evaluations.filter(evaluation_date__gte=date_from)
        if date_to:
            evaluations = evaluations.filter(evaluation_date__lte=date_to)
        
        # Get students for filter dropdown
        students = Profile.objects.filter(user_type='student').order_by('user__first_name', 'user__last_name')
        
        context = {
            'evaluations': evaluations,
            'students': students,
            'filters': {
                'student': student_filter,
                'severity': severity_filter,
                'hearing': hearing_filter,
                'date_from': date_from,
                'date_to': date_to,
            }
        }
        return render(request, 'app/evaluation_list.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')

@login_required
def evaluation_detail(request, evaluation_id):
    """View detailed evaluation"""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('student_dashboard')
        
        evaluation = get_object_or_404(StudentEvaluation, id=evaluation_id, is_draft=False)
        
        # Get student's evaluation history
        student_evaluations = StudentEvaluation.objects.filter(
            student=evaluation.student,
            is_draft=False
        ).exclude(id=evaluation_id).order_by('-evaluation_date')[:5]
        
        context = {
            'evaluation': evaluation,
            'student_evaluations': student_evaluations,
        }
        return render(request, 'app/evaluation_detail.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')

@login_required
def generate_evaluation_report(request):
    """Generate evaluation report with statistics"""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            return JsonResponse({'success': False, 'error': 'Permission denied'})
        
        # Get date range (default to last 30 days)
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=30)
        
        if request.GET.get('start_date'):
            start_date = datetime.strptime(request.GET.get('start_date'), '%Y-%m-%d').date()
        if request.GET.get('end_date'):
            end_date = datetime.strptime(request.GET.get('end_date'), '%Y-%m-%d').date()
        
        # Get evaluations in date range
        evaluations = StudentEvaluation.objects.filter(
            evaluation_date__range=[start_date, end_date],
            is_draft=False
        )
        
        # Calculate statistics
        total_evaluations = evaluations.count()
        severity_stats = evaluations.values('severity_level').annotate(count=Count('id'))
        reason_stats = evaluations.values('reason_for_session').annotate(count=Count('id'))
        hearing_stats = evaluations.values('hearing_frequency').annotate(count=Count('id'))
        
        # Most frequent students
        student_stats = evaluations.values(
            'student__user__first_name', 
            'student__user__last_name'
        ).annotate(count=Count('id')).order_by('-count')[:5]
        
        report_data = {
            'total_evaluations': total_evaluations,
            'date_range': {
                'start': start_date.strftime('%Y-%m-%d'),
                'end': end_date.strftime('%Y-%m-%d')
            },
            'severity_breakdown': list(severity_stats),
            'reason_breakdown': list(reason_stats),
            'hearing_breakdown': list(hearing_stats),
            'top_students': list(student_stats)
        }
        
        # Save report
        report = EvaluationReport.objects.create(
            title=f"Evaluation Report ({start_date} to {end_date})",
            generated_by=profile,
            date_range_start=start_date,
            date_range_end=end_date,
            report_data=report_data
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Report generated successfully!',
            'report_id': report.id,
            'report_data': report_data
        })
        
    except Profile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Profile not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})




@login_required
def student_settings(request):
    profile = Profile.objects.get(user=request.user)
    
    if request.method == 'POST':
        form_type = request.POST.get('form_type')
        
        if form_type == 'profile':
            user = request.user
            user.first_name = request.POST.get('first_name', '')
            user.last_name = request.POST.get('last_name', '')
            user.email = request.POST.get('email', '')
            user.save()
            
            profile.course = request.POST.get('course', '')
            profile.section = request.POST.get('section', '')
            
            if 'profile_picture' in request.FILES:
                profile.profile_picture = request.FILES['profile_picture']
            
            profile.save()
            messages.success(request, "Profile updated successfully!")
            
        elif form_type == 'password':
            from django.contrib.auth import update_session_auth_hash
            from django.contrib.auth.forms import PasswordChangeForm
            
            current_password = request.POST.get('current_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            
            if new_password != confirm_password:
                messages.error(request, "New passwords do not match.")
            elif not request.user.check_password(current_password):
                messages.error(request, "Incorrect current password.")
            else:
                request.user.set_password(new_password)
                request.user.save()
                update_session_auth_hash(request, request.user)
                messages.success(request, "Password updated successfully!")
        
        elif form_type == 'notifications':
            profile.email_notifications = request.POST.get('email_notifications') == 'on'
            profile.case_updates = request.POST.get('case_updates') == 'on'
            profile.session_reminders = request.POST.get('session_reminders') == 'on'
            profile.save()
            messages.success(request, "Notification preferences updated!")
            
        return redirect('student_settings')

    # Get user's first name for sidebar
    user_name = request.user.first_name or request.user.username
    # Get notifications
    notif_data = get_notifications(request)
    context = {
        'profile': profile,
        'user_name': user_name,
        'notifications': notif_data['notifications'],
        'unread_count': notif_data['unread_count'],
    }
    
    return render(request, 'app/student_settings.html', context)









from django.core.mail import send_mail

def finalize_evaluation(request, evaluation_id):
    evaluation = StudentEvaluation.objects.get(id=evaluation_id)

    # Mark evaluation as finalized
    evaluation.is_draft = False
    evaluation.save()

    # Get student's email
    student_email = evaluation.student.user.email
    student_name = evaluation.student.user.get_full_name()

    # Send notification email
    send_mail(
        subject='Your Evaluation Has Been Reviewed',
        message=f'Dear {student_name},\n\nYour evaluation has been reviewed and finalized by the guidance office.\n\nRegards,\nStudent Services Team',
        from_email=None,  # uses DEFAULT_FROM_EMAIL
        recipient_list=[student_email],
        fail_silently=False,
    )

    return redirect('evaluation_dashboard')  # or wherever your admin goes



from django.core.mail import send_mail
from django.http import JsonResponse, HttpResponse, FileResponse

def send_test_email(request):
    if request.method == "POST":
        email = request.POST.get("email")  # or whatever field you used
        if email:
            send_mail(
                subject="Test Email",
                message="This is a test email from the Student Evaluation System.",
                from_email=None,
                recipient_list=[email],
                fail_silently=False,
            )
            return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': 'Invalid request'})


@login_required
def calendar_events(request):
    try:
        from app.models import Hearing
        
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            return JsonResponse([], safe=False)

        # Get all scheduled sessions - FIXED: using scheduled_date instead of incident_date
        sessions = GuidanceSession.objects.filter(
            Q(assigned_counselor=profile) | Q(preferred_counselor=profile),
            scheduled_date__isnull=False
        ).select_related('student') # Optimization: fetch student data in one go

        # Get scheduled cases
        cases = Case.objects.filter(
            counselor=profile,
            scheduled_date__isnull=False
        ).select_related('student') # Optimization: fetch student data in one go
        
        # Get all hearings where user is involved
        hearings = Hearing.objects.filter(
            Q(presiding_officer=profile) | 
            Q(respondent=profile) |
            Q(complainant=profile) |
            Q(created_by=profile)
        ).select_related('respondent', 'case')

        events = []

        # Add sessions to events
        for session in sessions:
            # Get student name safely
            student_name = session.student.user.get_full_name() if session.student else "Unknown Student"
            
            # Create a descriptive title
            display_title = f"Session: {student_name} ({session.get_reason_display()})"
            
            # FIXED: Use scheduled_date instead of incident_date
            start_val = session.scheduled_date.isoformat()
            if session.scheduled_time:
                start_val = datetime.combine(session.scheduled_date, session.scheduled_time).isoformat()

            events.append({
                'title': display_title,
                'start': start_val,
                'url': reverse('counselor_session_detail', args=[session.id]),
                'color': '#3b82f6',
                'extendedProps': {
                    'type': 'session',
                    'status': session.status,
                    'student': student_name # Passing extra data just in case
                }
            })
        
        # Add cases to events
        for case in cases:
            student_name = case.student.user.get_full_name() if case.student else "Unknown Student"
            display_title = f"Case: {student_name} - {case.title}"
            
            start_val = case.scheduled_date.isoformat()
            if case.scheduled_time:
                start_val = datetime.combine(case.scheduled_date, case.scheduled_time).isoformat()

            events.append({
                'title': display_title,
                'start': start_val,
                'url': reverse('counselor_case_detail', args=[case.id]),
                'color': '#10b981',
                'extendedProps': {
                    'type': 'case',
                    'status': case.status,
                    'student': student_name
                }
            })
        
        # Add hearings to events
        for hearing in hearings:
            # Get respondent name safely
            respondent_name = hearing.respondent.user.get_full_name() if hearing.respondent else "Unknown"
            
            display_title = f"HEARING: {hearing.title} - {respondent_name}"
            
            start_val = hearing.scheduled_date.isoformat()
            if hearing.scheduled_time:
                start_val = datetime.combine(hearing.scheduled_date, hearing.scheduled_time).isoformat()
            
            events.append({
                'title': display_title,
                'start': start_val,
                'url': reverse('counselor_hearing_detail', args=[hearing.id]),
                'color': '#f59e0b',  # Orange/yellow color for hearings
                'extendedProps': {
                    'type': 'hearing',
                    'hearing_id': hearing.id,
                    'hearing_number': hearing.hearing_number,
                    'status': hearing.status
                }
            })

        return JsonResponse(events, safe=False)
    except Exception as e:
        # Make sure 'logger' is imported at the top of your file
        # import logging; logger = logging.getLogger(__name__)
        print(f"Error: {e}") 
        return JsonResponse([], safe=False)


def verify_email(request, user_id):
    try:
        # Fetch the user object using the user_id from the URL
        user = get_object_or_404(User, id=user_id)
        
        if request.method == 'POST':
            otp_input = request.POST.get('otp')
            verification = EmailVerification.objects.get(user=user)

            if verification.otp == otp_input:
                # If OTP is correct, mark email as verified and activate the user
                verification.verified = True
                verification.save()

                user.is_active = True
                user.save()

                messages.success(request, "Email successfully verified! You can now log in.")
                return redirect('login')  # Redirect to the login page
            else:
                messages.error(request, "Invalid OTP. Please try again.")

        return render(request, 'app/verify_email.html')  # Render the verification template

    except EmailVerification.DoesNotExist:
        messages.error(request, "Verification record not found.")
        return redirect('home')
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('home')



# Case Management Views
@login_required
def edit_case(request):
    """Edit case details"""
    if request.method == 'POST':
        case_id = request.POST.get('case_id')
        title = request.POST.get('title')
        description = request.POST.get('description')
        
        print(f"DEBUG: Editing case_id={case_id}, title={title}, description={description[:50]}...")
        
        try:
            # Try to get Case
            case = Case.objects.get(id=case_id, student__user=request.user, is_active=True)
            print(f"DEBUG: Found Case - Old title: {case.title}")
            case.title = title
            case.description = description
            case.save()
            print(f"DEBUG: Case saved - New title: {case.title}")
            
            return JsonResponse({
                'success': True,
                'message': 'Case updated successfully!'
            })
        except Case.DoesNotExist:
            # Try to get GuidanceSession
            try:
                session = GuidanceSession.objects.get(id=case_id, student__user=request.user, is_active=True)
                print(f"DEBUG: Found Session - Old description: {session.concern_description[:50]}")
                session.concern_description = description
                session.save()
                print(f"DEBUG: Session saved - New description: {session.concern_description[:50]}")
                
                return JsonResponse({
                    'success': True,
                    'message': 'Session updated successfully!'
                })
            except GuidanceSession.DoesNotExist:
                print(f"DEBUG: Case/Session not found for id={case_id}, user={request.user}")
                return JsonResponse({
                    'success': False,
                    'message': 'Case or session not found.'
                }, status=404)
        except Exception as e:
            print(f"DEBUG: Error - {str(e)}")
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method.'
    }, status=400)

@login_required
def cancel_case(request):
    """Cancel a case or session"""
    if request.method == 'POST':
        case_id = request.POST.get('case_id')
        case_type = request.POST.get('case_type', 'case')
        
        try:
            if case_type == 'case':
                case = Case.objects.get(id=case_id, student__user=request.user)
                case.is_active = False
                case.status = 'cancelled'
                case.save()
                message = 'Case cancelled successfully!'
            else:
                session = GuidanceSession.objects.get(id=case_id, student__user=request.user)
                session.is_active = False
                session.status = 'canceled'
                session.save()
                message = 'Session cancelled successfully!'
            
            return JsonResponse({
                'success': True,
                'message': message
            })
        except (Case.DoesNotExist, GuidanceSession.DoesNotExist):
            return JsonResponse({
                'success': False,
                'message': 'Case or session not found.'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method.'
    }, status=400)

@login_required
def student_cases_only(request):
    """Display only formal Cases (not sessions)"""
    if hasattr(request.user, 'profile'):
        student_profile = request.user.profile
        user_name = request.user.first_name or request.user.username
        
        # Get only Cases (not sessions)
        all_cases = Case.objects.filter(student=student_profile, is_active=True).order_by('-created_at')
        
        # Statistics
        pending_cases = all_cases.filter(status='pending')
        ongoing_cases = all_cases.filter(status__in=['in_progress', 'under_review', 'approved'])
        completed_cases = all_cases.filter(status__in=['resolved', 'completed'])
        
        # Prepare case data with linked sessions count
        cases_data = []
        for case in all_cases:
            linked_sessions = case.sessions.all()
            cases_data.append({
                'id': case.id,
                'title': case.title,
                'description': case.description,
                'status': case.status,
                'counselor': case.counselor,
                'created_at': case.created_at,
                'updated_at': case.updated_at,
                'session_count': linked_sessions.count(),
                'sessions': linked_sessions
            })
    # Get notifications
        notif_data = get_notifications(request)
        context = {
            'user_name': user_name,
            'cases': cases_data,
            'pending_count': pending_cases.count(),
            'ongoing_count': ongoing_cases.count(),
            'completed_count': completed_cases.count(),
            'total_count': all_cases.count(),
        }
        
        return render(request, 'app/student_cases_only.html', context)
    
    return redirect('login')

@login_required
def student_sessions_only(request):
    """Display only Guidance Sessions"""
    if hasattr(request.user, 'profile'):
        student_profile = request.user.profile
        user_name = request.user.first_name or request.user.username
        
        # Get all sessions
        all_sessions = GuidanceSession.objects.filter(student=student_profile, is_active=True).order_by('-created_at')
        
        # Statistics
        pending_sessions = all_sessions.filter(status='pending')
        ongoing_sessions = all_sessions.filter(status__in=['approved', 'in_progress'])
        completed_sessions = all_sessions.filter(status__in=['completed'])
        
        # Prepare session data with linked case info
        sessions_data = []
        for session in all_sessions:
            linked_case = Case.objects.filter(sessions=session).first()
            sessions_data.append({
                'id': session.id,
                'reason': session.get_reason_display(),
                'concern_description': session.concern_description,
                'status': session.status,
                'counselor': session.assigned_counselor or session.preferred_counselor,
                'scheduled_date': session.scheduled_date,
                'scheduled_time': session.scheduled_time,
                'created_at': session.created_at,
                'linked_case': linked_case,
                'is_linked': linked_case is not None
            })
    # Get notifications
        notif_data = get_notifications(request)
        context = {
            'user_name': user_name,
            'sessions': sessions_data,
            'pending_count': pending_sessions.count(),
            'ongoing_count': ongoing_sessions.count(),
            'completed_count': completed_sessions.count(),
            'total_count': all_sessions.count(),
        }
        
        return render(request, 'app/student_sessions_only.html', context)
    
    return redirect('login')


@login_required
def case_detail_view(request, case_id):
    """Detailed view for a specific case"""
    try:
        profile = Profile.objects.get(user=request.user)
        case = get_object_or_404(Case, id=case_id)
        
        # Check permission: student can only view their own cases, counselors can view all
        if profile.user_type == 'student' and case.student != profile:
            messages.error(request, "You don't have permission to view this case.")
            return redirect('student_cases_only')
        
        # Get all sessions linked to this case
        linked_sessions = case.sessions.all().order_by('-created_at')
        
        context = {
            'profile': profile,
            'user_name': request.user.first_name if request.user.first_name else request.user.username,
            'case': case,
            'linked_sessions': linked_sessions,
            'is_student': profile.user_type == 'student',
        }
        return render(request, 'app/case_detail_view.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')


@login_required
def session_detail_view(request, session_id):
    """Detailed view for a specific guidance session"""
    try:
        profile = Profile.objects.get(user=request.user)
        session = get_object_or_404(GuidanceSession, id=session_id)
        
        # Check permission: student can only view their own sessions, counselors can view all
        if profile.user_type == 'student' and session.student != profile:
            messages.error(request, "You don't have permission to view this session.")
            return redirect('student_sessions_only')
        
        # Get all cases this session is linked to
        linked_cases = session.cases.all().order_by('-created_at')
        
        # Check if session is editable (pending AND not scheduled)
        is_editable = (session.status == 'pending' and not session.scheduled_date)
        
        context = {
            'profile': profile,
            'user_name': request.user.first_name if request.user.first_name else request.user.username,
            'session': session,
            'linked_cases': linked_cases,
            'is_student': profile.user_type == 'student',
            'is_editable': is_editable,
        }
        return render(request, 'app/session_detail_view.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')


# ============================================
# Counselor Dashboard Views
# ============================================

@login_required
def counselor_dashboard(request):
    """Main counselor dashboard view"""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('dashboard')
        
        # Get all sessions
        all_sessions = GuidanceSession.objects.all().order_by('-created_at')
        
        # Count statistics
        pending_sessions_count = all_sessions.filter(status='pending').count()
        active_cases_count = Case.objects.filter(status='in_progress', is_active=True).count()
        
        # Completed this month
        from datetime import datetime, timedelta
        first_day = datetime.now().replace(day=1)
        completed_count = all_sessions.filter(
            status='completed',
            updated_at__gte=first_day
        ).count()
        
        # Calculate percentage (simplified)
        completed_percentage = 15  # Placeholder
        
        total_students = Profile.objects.filter(user_type='student').count()
        
        # Pagination setup
        from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
        
        # Pagination for Pending Sessions
        pending_sessions_list = all_sessions.filter(status='pending')
        pending_page = request.GET.get('pending_page', 1)
        pending_paginator = Paginator(pending_sessions_list, 10)  # 10 items per page
        try:
            pending_sessions = pending_paginator.page(pending_page)
        except PageNotAnInteger:
            pending_sessions = pending_paginator.page(1)
        except EmptyPage:
            pending_sessions = pending_paginator.page(pending_paginator.num_pages)
        
        # Pagination for Active Cases
        active_cases_list = Case.objects.filter(
            status__in=['pending', 'in_progress'],
            is_active=True
        ).select_related('student__user', 'counselor__user').order_by('-updated_at')
        cases_page = request.GET.get('cases_page', 1)
        cases_paginator = Paginator(active_cases_list, 10)
        try:
            active_cases = cases_paginator.page(cases_page)
        except PageNotAnInteger:
            active_cases = cases_paginator.page(1)
        except EmptyPage:
            active_cases = cases_paginator.page(cases_paginator.num_pages)
        
        # Pagination for Upcoming Schedules
        upcoming_schedules_list = all_sessions.filter(
            status='approved',
            scheduled_date__gte=datetime.now().date()
        ).select_related('student__user').order_by('scheduled_date', 'scheduled_time')
        schedules_page = request.GET.get('schedules_page', 1)
        schedules_paginator = Paginator(upcoming_schedules_list, 10)
        try:
            upcoming_schedules = schedules_paginator.page(schedules_page)
        except PageNotAnInteger:
            upcoming_schedules = schedules_paginator.page(1)
        except EmptyPage:
            upcoming_schedules = schedules_paginator.page(schedules_paginator.num_pages)
        
        context = {
            'profile': profile,
            'user_name': request.user.get_full_name() or request.user.username,
            'pending_sessions_count': pending_sessions_count,
            'active_cases_count': active_cases_count,
            'completed_count': completed_count,
            'completed_percentage': completed_percentage,
            'total_students': total_students,
            'pending_sessions': pending_sessions,
            'active_cases': active_cases,
            'upcoming_schedules': upcoming_schedules,
        }
        return render(request, 'app/counselor_dashboard.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')


@login_required
def counselor_sessions(request):
    """Session management view with tabs"""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('dashboard')
        
        # Get all sessions
        all_sessions = GuidanceSession.objects.all().select_related(
            'student__user', 'preferred_counselor__user', 'assigned_counselor__user'
        ).order_by('-created_at')
        
        # Filter by status
        pending_sessions = all_sessions.filter(status='pending')
        scheduled_sessions = all_sessions.filter(status='approved')
        completed_sessions = all_sessions.filter(status='completed')
        
        # Count statistics
        pending_count = pending_sessions.count()
        scheduled_count = scheduled_sessions.count()
        completed_count = completed_sessions.count()
        total_count = all_sessions.count()
        upcoming_week_count = scheduled_sessions.filter(
            scheduled_date__gte=timezone.now().date(),
            scheduled_date__lte=timezone.now().date() + timedelta(days=7)
        ).count()
        group_sessions_count = all_sessions.filter(participants__isnull=False).distinct().count()
        
        context = {
            'profile': profile,
            'user_name': request.user.get_full_name() or request.user.username,
            'pending_sessions_count': pending_count,
            'pending_sessions': pending_sessions,
            'scheduled_sessions': scheduled_sessions,
            'completed_sessions': completed_sessions,
            'all_sessions': all_sessions,
            'pending_count': pending_count,
            'scheduled_count': scheduled_count,
            'completed_count': completed_count,
            'total_count': total_count,
            'upcoming_week_count': upcoming_week_count,
            'group_sessions_count': group_sessions_count,
        }
        return render(request, 'app/counselor_sessions.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')


@login_required
def counselor_session_detail(request, session_id):
    """Detailed view of a specific session"""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('dashboard')
        
        session = get_object_or_404(
            GuidanceSession.objects.select_related(
                'student__user', 'preferred_counselor__user', 'assigned_counselor__user'
            ),
            id=session_id
        )
        
        from datetime import date
        today = date.today()
        
        context = {
            'profile': profile,
            'user_name': request.user.get_full_name() or request.user.username,
            'pending_sessions_count': GuidanceSession.objects.filter(status='pending').count(),
            'session': session,
            'today': today,
        }
        return render(request, 'app/counselor_session_detail.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')


@login_required
def counselor_approve_session(request, session_id):
    """Approve and schedule a session"""
    if request.method == 'POST':
        try:
            from app.utils import NotificationManager
            
            profile = Profile.objects.get(user=request.user)
            if profile.user_type != 'counselor':
                messages.error(request, "You don't have permission to perform this action.")
                return redirect('dashboard')
            
            session = get_object_or_404(GuidanceSession, id=session_id)
            
            # Update session
            session.status = 'approved'
            session.assigned_counselor = profile
            session.scheduled_date = request.POST.get('scheduled_date')
            session.scheduled_time = request.POST.get('scheduled_time')
            session.save()
            
            # Send notifications using NotificationManager
            NotificationManager.notify_session_approved(session)
            NotificationManager.notify_session_scheduled(session)
            
            messages.success(request, 'Session approved and scheduled successfully!')
            return redirect('counselor_session_detail', session_id=session.id)
        except Exception as e:
            messages.error(request, f'Error approving session: {str(e)}')
            return redirect('counselor_session_detail', session_id=session_id)
    
    return redirect('counselor_sessions')


@login_required
def counselor_complete_session(request, session_id):
    """Mark a session as completed"""
    if request.method == 'POST':
        try:
            from app.utils import NotificationManager
            
            profile = Profile.objects.get(user=request.user)
            if profile.user_type != 'counselor':
                messages.error(request, "You don't have permission to perform this action.")
                return redirect('dashboard')
            
            session = get_object_or_404(GuidanceSession, id=session_id)
            
            # Update session
            session.status = 'completed'
            session.session_notes = request.POST.get('session_notes')
            session.save()
            
            # Send notification using NotificationManager
            NotificationManager.notify_session_completed(session)
            
            messages.success(request, 'Session marked as completed!')
            return redirect('counselor_session_detail', session_id=session.id)
        except Exception as e:
            messages.error(request, f'Error completing session: {str(e)}')
            return redirect('counselor_session_detail', session_id=session_id)
    
    return redirect('counselor_sessions')


@login_required
def counselor_reject_session(request, session_id):
    """Reject a session request"""
    if request.method == 'POST':
        try:
            from app.utils import NotificationManager
            
            profile = Profile.objects.get(user=request.user)
            if profile.user_type != 'counselor':
                messages.error(request, "You don't have permission to perform this action.")
                return redirect('dashboard')
            
            session = get_object_or_404(GuidanceSession, id=session_id)
            
            # Update session
            session.status = 'cancelled'
            rejection_reason = request.POST.get('rejection_reason', 'No reason provided')
            session.save()
            
            # Send notification using NotificationManager
            NotificationManager.notify_session_rejected(session, rejection_reason)
            
            messages.success(request, 'Session request rejected.')
            return redirect('counselor_sessions')
        except Exception as e:
            messages.error(request, f'Error rejecting session: {str(e)}')
            return redirect('counselor_session_detail', session_id=session_id)
    
    return redirect('counselor_sessions')


# Case Management Views
@login_required
def counselor_cases(request):
    """Case management view"""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('dashboard')
        
        db_cases = Case.objects.filter(is_active=True).select_related(
            'student__user', 'counselor__user'
        ).order_by('-created_at')

        # Include standalone guidance sessions not yet linked to a Case.
        standalone_sessions = GuidanceSession.objects.filter(case__isnull=True).select_related(
            'student__user', 'assigned_counselor__user'
        ).order_by('-created_at')

        def normalize_status(record_status, is_session=False):
            if is_session:
                if record_status == 'approved':
                    return 'in_progress'
                if record_status == 'completed':
                    return 'completed'
                if record_status == 'canceled':
                    return 'cancelled'
            return record_status

        all_cases = []
        for c in db_cases:
            age_days = max((timezone.now() - c.created_at).days, 0)
            all_cases.append({
                'id': c.id,
                'record_type': 'case',
                'source_label': 'Formal Case',
                'title': c.title,
                'description': c.description,
                'student': c.student,
                'counselor': c.counselor,
                'status': normalize_status(c.status),
                'sessions_count': c.sessions.count(),
                'age_days': age_days,
                'created_at': c.created_at,
                'updated_at': c.updated_at,
            })

        for s in standalone_sessions:
            age_days = max((timezone.now() - s.created_at).days, 0)
            all_cases.append({
                'id': s.id,
                'record_type': 'session',
                'source_label': 'Session Intake',
                'title': s.title or f"Guidance Session #{s.id}",
                'description': s.concern_description or '',
                'student': s.student,
                'counselor': s.assigned_counselor,
                'status': normalize_status(s.status, is_session=True),
                'sessions_count': 1,
                'age_days': age_days,
                'created_at': s.created_at,
                'updated_at': s.updated_at,
            })

        all_cases.sort(key=lambda x: x['created_at'], reverse=True)

        active_cases = [c for c in all_cases if c['status'] in ['pending', 'in_progress']]
        resolved_cases = [c for c in all_cases if c['status'] == 'completed']
        overdue_count = sum(1 for c in active_cases if c['age_days'] > 14)

        pending_count = sum(1 for c in all_cases if c['status'] == 'pending')
        in_progress_count = sum(1 for c in all_cases if c['status'] == 'in_progress')
        resolved_count = len(resolved_cases)
        total_count = len(all_cases)
        active_count = len(active_cases)

        # Case aging buckets for lifecycle chart
        aging_labels = ['0-7 Days', '8-14 Days', '15+ Days']
        aging_values = [
            sum(1 for c in active_cases if c['age_days'] <= 7),
            sum(1 for c in active_cases if 8 <= c['age_days'] <= 14),
            sum(1 for c in active_cases if c['age_days'] > 14),
        ]
        
        context = {
            'profile': profile,
            'user_name': request.user.get_full_name() or request.user.username,
            'pending_sessions_count': GuidanceSession.objects.filter(status='pending').count(),
            'all_cases': all_cases,
            'active_cases': active_cases,
            'resolved_cases': resolved_cases,
            'pending_count': pending_count,
            'in_progress_count': in_progress_count,
            'resolved_count': resolved_count,
            'total_count': total_count,
            'active_count': active_count,
            'overdue_count': overdue_count,
            'case_aging_labels': json.dumps(aging_labels),
            'case_aging_values': json.dumps(aging_values),
        }
        return render(request, 'app/counselor_cases.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')


@login_required
def counselor_case_detail(request, case_id):
    """Detailed view of a specific case. Supports POST to update status."""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('dashboard')
        
        case = get_object_or_404(
            Case.objects.select_related('student__user', 'counselor__user'),
            id=case_id
        )
        
        if request.method == 'POST':
            new_status = request.POST.get('status')
            if new_status in ['pending', 'in_progress', 'completed']:
                case.status = new_status
                case.save()
                # Notify student
                try:
                    Notification.objects.create(
                        recipient=case.student,
                        sender=profile,
                        notification_type='case_updated',
                        title='Case status updated',
                        message=f'Case "{case.title}" status changed to {new_status.replace("_"," ")}.',
                        link=f'/case/{case.id}/detail/'
                    )
                except Exception:
                    pass
                messages.success(request, 'Case status updated.')
                return redirect('counselor_case_detail', case_id=case.id)
        
        linked_sessions = case.sessions.all().order_by('-created_at')
        
        # For linking modal: unlinked sessions for this student
        available_sessions = GuidanceSession.objects.filter(student=case.student).exclude(cases__id=case.id).order_by('-created_at')[:50]
        
        context = {
            'profile': profile,
            'user_name': request.user.get_full_name() or request.user.username,
            'pending_sessions_count': GuidanceSession.objects.filter(status='pending').count(),
            'case': case,
            'linked_sessions': linked_sessions,
            'available_sessions': available_sessions,
        }
        return render(request, 'app/counselor_case_detail.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')


@login_required
def counselor_create_case(request):
    """Create a new case"""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('dashboard')
        
        if request.method == 'POST':
            student_id = request.POST.get('student')
            title = request.POST.get('title')
            description = request.POST.get('description')
            status = request.POST.get('status', 'pending')
            session_id = request.POST.get('session_id')
            
            student = get_object_or_404(Profile, id=student_id, user_type='student')
            
            case = Case.objects.create(
                student=student,
                counselor=profile,
                title=title,
                description=description,
                status=status,
                is_active=True
            )
            
            # Link session if provided
            if session_id:
                session = GuidanceSession.objects.get(id=session_id)
                case.sessions.add(session)
            
            # Send notification using NotificationManager
            from app.utils import NotificationManager
            NotificationManager.create_notification(
                'case_created',
                recipient=student,
                sender=profile,
                link=reverse('view_case_details', args=[case.id]),
                student=student.user.get_full_name(),
                title=title
            )
            
            messages.success(request, 'Case created successfully!')
            return redirect('counselor_case_detail', case_id=case.id)
        
        # Prefill from query params
        selected_student_id = request.GET.get('student')
        selected_session_id = request.GET.get('session_id')

        students = Profile.objects.filter(user_type='student').select_related('user').order_by('user__first_name')
        recent_sessions = GuidanceSession.objects.all().select_related('student__user').order_by('-created_at')[:50]

        # Ensure the selected session appears in the list even if older than window
        prefill_title = None
        prefill_description = None
        if selected_session_id:
            try:
                sel_session = GuidanceSession.objects.select_related('student__user').get(id=selected_session_id)
                if sel_session not in recent_sessions:
                    recent_sessions = list(recent_sessions)
                    recent_sessions.insert(0, sel_session)
                # Build sensible defaults
                reason_label = sel_session.get_reason_display() if hasattr(sel_session, 'get_reason_display') else str(sel_session.reason)
                # Generate a unique placeholder case title (editable)
                prefill_title = f"Case {timezone.now().strftime('%Y%m%d')}-{get_random_string(4).upper()}"
                base_desc = sel_session.concern_description if hasattr(sel_session, 'concern_description') else ''
                date_str = sel_session.created_at.strftime('%b %d, %Y') if hasattr(sel_session, 'created_at') and sel_session.created_at else ''
                prefill_description = (
                    f"Origin: Session #{sel_session.id} ({date_str})\n"
                    f"Reason: {reason_label}\n\n"
                    f"Student concern (copied):\n{base_desc}\n\n"
                    f"Additional notes:"
                )
            except GuidanceSession.DoesNotExist:
                selected_session_id = None

        context = {
            'profile': profile,
            'user_name': request.user.get_full_name() or request.user.username,
            'pending_sessions_count': GuidanceSession.objects.filter(status='pending').count(),
            'students': students,
            'recent_sessions': recent_sessions,
            'selected_student_id': int(selected_student_id) if selected_student_id else None,
            'selected_session_id': int(selected_session_id) if selected_session_id else None,
            'prefill_title': prefill_title,
            'prefill_description': prefill_description,
        }
        return render(request, 'app/counselor_create_case.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')


# Placeholder views for remaining pages
@login_required
def counselor_students(request):
    profile = Profile.objects.get(user=request.user)
    from django.db.models import Count
    students = (
        Profile.objects.filter(user_type='student')
        .select_related('user')
        .annotate(cases_count=Count('cases', distinct=True), sessions_count=Count('sessions', distinct=True))
        .order_by('user__first_name')
    )
    context = {
        'profile': profile,
        'user_name': request.user.get_full_name() or request.user.username,
        'pending_sessions_count': GuidanceSession.objects.filter(status='pending').count(),
        'students': students,
    }
    return render(request, 'app/counselor_students.html', context)


@login_required
def counselor_student_profile(request, student_id):
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('dashboard')
        student = get_object_or_404(Profile.objects.select_related('user'), id=student_id, user_type='student')
        sessions_qs = GuidanceSession.objects.filter(student=student).select_related('student__user','assigned_counselor__user','preferred_counselor__user').order_by('-created_at')
        cases_qs = Case.objects.filter(student=student).select_related('student__user','counselor__user').order_by('-created_at')
        completed_sessions = sessions_qs.filter(status='completed').count()
        notif_data = get_notifications(request)
        context = {
            'profile': profile,
            'user_name': request.user.get_full_name() or request.user.username,
            'pending_sessions_count': GuidanceSession.objects.filter(status='pending').count(),
            'student': student,
            'sessions': sessions_qs,
            'cases': cases_qs,
            'sessions_count': sessions_qs.count(),
            'cases_count': cases_qs.count(),
            'completed_sessions': completed_sessions,
            'notifications': notif_data['notifications'],
            'unread_count': notif_data['unread_count'],
        }
        return render(request, 'app/counselor_student_profile.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')


@login_required
def counselor_calendar(request):
    profile = Profile.objects.get(user=request.user)
    context = {
        'profile': profile,
        'user_name': request.user.get_full_name() or request.user.username,
        'pending_sessions_count': GuidanceSession.objects.filter(status='pending').count(),
    }
    return render(request, 'app/counselor_calendar.html', context)


@login_required
def counselor_reports(request):
    profile = Profile.objects.get(user=request.user)
    context = {
        'profile': profile,
        'user_name': request.user.get_full_name() or request.user.username,
        'pending_sessions_count': GuidanceSession.objects.filter(status='pending').count(),
    }
    return render(request, 'app/counselor_reports.html', context)


@login_required
def counselor_settings(request):
    profile = Profile.objects.get(user=request.user)
    context = {
        'profile': profile,
        'user_name': request.user.get_full_name() or request.user.username,
        'pending_sessions_count': GuidanceSession.objects.filter(status='pending').count(),
    }
    return render(request, 'app/counselor_settings.html', context)


# Notification Views
@login_required
def get_notifications(request):
    """Get notifications for current user"""
    try:
        profile = Profile.objects.get(user=request.user)
        notifications = Notification.objects.filter(recipient=profile).order_by('-created_at')[:10]
        unread_count = notifications.filter(is_read=False).count()
        
        return {
            'notifications': notifications,
            'unread_count': unread_count
        }
    except:
        return {'notifications': [], 'unread_count': 0}


@login_required
def mark_notification_read(request, notif_id):
    """Mark a single notification as read"""
    if request.method == 'POST':
        try:
            profile = Profile.objects.get(user=request.user)
            notif = Notification.objects.get(id=notif_id, recipient=profile)
            notif.is_read = True
            notif.save()
            return JsonResponse({'success': True})
        except:
            return JsonResponse({'success': False}, status=400)
    return JsonResponse({'success': False}, status=405)


@login_required
def mark_all_notifications_read(request):
    """Mark all notifications as read"""
    if request.method == 'POST':
        try:
            profile = Profile.objects.get(user=request.user)
            Notification.objects.filter(recipient=profile, is_read=False).update(is_read=True)
            return JsonResponse({'success': True})
        except:
            return JsonResponse({'success': False}, status=400)
    return JsonResponse({'success': False}, status=405)


@login_required
def view_all_notifications(request):
    """View all notifications page"""
    try:
        profile = Profile.objects.get(user=request.user)
        notifications = Notification.objects.filter(recipient=profile).order_by('-created_at')
        unread_count = notifications.filter(is_read=False).count()
        
        context = {
            'profile': profile,
            'user_name': request.user.get_full_name() or request.user.username,
            'notifications': notifications,
            'unread_count': unread_count,
            'pending_sessions_count': GuidanceSession.objects.filter(status='pending').count() if profile.user_type == 'counselor' else 0,
        }
        return render(request, 'app/notifications.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')


# Update student profile view with complete data
def counselor_student_profile(request, student_id):
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('dashboard')
        
        student = get_object_or_404(
            Profile.objects.select_related('user'),
            id=student_id,
            user_type='student'
        )
        
        completed_sessions = student.sessions.filter(status='completed').count()
    # Get notifications
        notif_data = get_notifications(request)
        context = {
            'profile': profile,
            'user_name': request.user.get_full_name() or request.user.username,
            'pending_sessions_count': GuidanceSession.objects.filter(status='pending').count(),
            'student': student,
            'completed_sessions': completed_sessions,
            'notifications': notif_data['notifications'],
            'unread_count': notif_data['unread_count'],
        }
        return render(request, 'app/counselor_student_profile.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')


# Calendar View with data
def counselor_calendar(request):
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('dashboard')
        
        from datetime import date, timedelta
        today = date.today()
        
        upcoming_sessions = GuidanceSession.objects.filter(
            status='approved',
            scheduled_date__gte=today
        ).select_related('student__user').order_by('scheduled_date', 'scheduled_time')[:20]
        
        today_sessions = GuidanceSession.objects.filter(
            status='approved',
            scheduled_date=today
        ).select_related('student__user').order_by('scheduled_time')
        notif_data = get_notifications(request)
        context = {
            'profile': profile,
            'user_name': request.user.get_full_name() or request.user.username,
            'pending_sessions_count': GuidanceSession.objects.filter(status='pending').count(),
            'upcoming_sessions': upcoming_sessions,
            'today_sessions': today_sessions,
            'notifications': notif_data['notifications'],
            'unread_count': notif_data['unread_count'],
        }
        return render(request, 'app/counselor_calendar.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')


# Reports View with analytics
def counselor_reports(request):
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('dashboard')
        
        import json
        from django.db.models.functions import TruncMonth
        from django.utils import timezone
        
        # Statistics
        total_sessions = GuidanceSession.objects.count()
        base_cases_qs = Case.objects.filter(is_active=True)
        # Include standalone sessions that do not yet have a linked Case record.
        standalone_sessions_qs = GuidanceSession.objects.filter(case__isnull=True)
        total_cases = base_cases_qs.count() + standalone_sessions_qs.count()
        total_students = Profile.objects.filter(user_type='student').count()
        completed_sessions = GuidanceSession.objects.filter(status='completed').count()
        completion_rate = int((completed_sessions / total_sessions * 100)) if total_sessions > 0 else 0
        
        # Sessions by month (real data for current year)
        current_year = timezone.now().year
        month_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                        'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        month_counts = {i: 0 for i in range(1, 13)}
        monthly_rows = (
            GuidanceSession.objects
            .filter(created_at__year=current_year)
            .annotate(month=TruncMonth('created_at'))
            .values('month')
            .annotate(total=Count('id'))
            .order_by('month')
        )
        for row in monthly_rows:
            if row['month']:
                month_counts[row['month'].month] = row['total']
        sessions_by_month = [month_counts[i] for i in range(1, 13)]
        
        # Cases by status
        pending_cases = (
            base_cases_qs.filter(status='pending').count()
            + standalone_sessions_qs.filter(status='pending').count()
        )
        in_progress_cases = (
            base_cases_qs.filter(status='in_progress').count()
            + standalone_sessions_qs.filter(status='approved').count()
        )
        resolved_cases = (
            base_cases_qs.filter(status='completed').count()
            + standalone_sessions_qs.filter(status='completed').count()
        )
        cases_by_status = [pending_cases, in_progress_cases, resolved_cases]
        
        # Session reasons
        reason_data = GuidanceSession.objects.values('reason').annotate(count=Count('reason'))
        reason_labels = [dict(GuidanceSession.REASON_CHOICES).get(r['reason'], r['reason']) for r in reason_data]
        reason_counts = [r['count'] for r in reason_data]
        notif_data = get_notifications(request)
        context = {
            'profile': profile,
            'user_name': request.user.get_full_name() or request.user.username,
            'pending_sessions_count': GuidanceSession.objects.filter(status='pending').count(),
            'total_sessions': total_sessions,
            'total_cases': total_cases,
            'total_students': total_students,
            'completion_rate': completion_rate,
            'month_labels': json.dumps(month_labels),
            'sessions_by_month': json.dumps(sessions_by_month),
            'cases_by_status': json.dumps(cases_by_status),
            'reason_labels': json.dumps(reason_labels),
            'reason_counts': json.dumps(reason_counts),
            'notifications': notif_data['notifications'],
            'unread_count': notif_data['unread_count'],
        }
        return render(request, 'app/counselor_reports.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Your profile was not found")
        return redirect('home')

@login_required
def counselor_reschedule_session(request, session_id):
    """Reschedule an approved session"""
    if request.method == 'POST':
        try:
            profile = Profile.objects.get(user=request.user)
            if profile.user_type != 'counselor':
                messages.error(request, "You don't have permission to perform this action.")
                return redirect('dashboard')

            session = get_object_or_404(GuidanceSession, id=session_id)
            session.scheduled_date = request.POST.get('scheduled_date')
            session.scheduled_time = request.POST.get('scheduled_time')
            session.save()

            Notification.objects.create(
                recipient=session.student,
                sender=profile,
                notification_type='session_scheduled',
                title='Session Rescheduled',
                message=f'Your session has been rescheduled to {session.scheduled_date}.',
                link=f'/session/{session.id}/detail/'
            )

            messages.success(request, 'Session rescheduled successfully!')
            return redirect('counselor_session_detail', session_id=session.id)
        except Exception as e:
            messages.error(request, f'Error rescheduling session: {str(e)}')
            return redirect('counselor_session_detail', session_id=session_id)
    return redirect('counselor_sessions')


@login_required
def counselor_link_session_to_case(request, case_id):
    """Link a session to a case"""
    if request.method == 'POST':
        try:
            profile = Profile.objects.get(user=request.user)
            if profile.user_type != 'counselor':
                messages.error(request, "You don't have permission to perform this action.")
                return redirect('dashboard')

            case = get_object_or_404(Case, id=case_id)
            session_id = request.POST.get('session_id')
            session = get_object_or_404(GuidanceSession, id=session_id)
            case.sessions.add(session)

            Notification.objects.create(
                recipient=case.student,
                sender=profile,
                notification_type='case_updated',
                title='Session Linked to Case',
                message=f'Session #{session.id} has been linked to your case "{case.title}".',
                link=f'/case/{case.id}/detail/'
            )

            messages.success(request, 'Session linked to case.')
            return redirect('counselor_case_detail', case_id=case.id)
        except Exception as e:
            messages.error(request, f'Error linking session: {str(e)}')
            return redirect('counselor_case_detail', case_id=case_id)
    return redirect('counselor_cases')


@login_required
def counselor_reports_export(request, format):
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            return HttpResponse(status=403)
        from django.db.models import Count
        sessions = GuidanceSession.objects.all().select_related('student__user')
        cases = Case.objects.filter(is_active=True).select_related('student__user', 'counselor__user')
        if format == 'excel':
            import io
            from django.db.models.functions import TruncMonth
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws_summary = wb.active
            ws_summary.title = 'Summary Report'

            # Shared styles
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            thin_border = Border(
                left=Side(style='thin', color='BFBFBF'),
                right=Side(style='thin', color='BFBFBF'),
                top=Side(style='thin', color='BFBFBF'),
                bottom=Side(style='thin', color='BFBFBF'),
            )

            def style_sheet(sheet):
                # Header row formatting
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                # Data cell borders
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row,
                                           min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical='center')

                # Enable filter + freeze header
                sheet.auto_filter.ref = sheet.dimensions
                sheet.freeze_panes = 'A2'

                # Auto-fit columns by content length
                for column in sheet.columns:
                    max_len = 0
                    col_letter = column[0].column_letter
                    for cell in column:
                        val = '' if cell.value is None else str(cell.value)
                        if len(val) > max_len:
                            max_len = len(val)
                    sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

            # ---------- Sheet 1: Summary Report ----------
            total_sessions = GuidanceSession.objects.count()
            total_students = Profile.objects.filter(user_type='student').count()
            total_cases = Case.objects.filter(is_active=True).count()
            pending_cases = Case.objects.filter(status='pending', is_active=True).count()
            in_progress_cases = Case.objects.filter(status='in_progress', is_active=True).count()
            resolved_cases = Case.objects.filter(status='completed', is_active=True).count()
            total_counselors = Profile.objects.filter(user_type='counselor').count()

            ws_summary.append(['Metric', 'Value'])
            ws_summary.append(['Total Counseling Sessions', total_sessions])
            ws_summary.append(['Total Students Assisted', total_students])
            ws_summary.append(['Total Cases Recorded', total_cases])
            ws_summary.append(['Pending Cases', pending_cases])
            ws_summary.append(['In Progress Cases', in_progress_cases])
            ws_summary.append(['Resolved Cases', resolved_cases])
            ws_summary.append(['Total Counselors', total_counselors])
            style_sheet(ws_summary)

            # ---------- Sheet 2: Counseling Sessions ----------
            ws_sessions = wb.create_sheet('Counseling Sessions')
            ws_sessions.append([
                'Session ID', 'Student ID', 'Student Name', 'Course', 'Year Level',
                'Counselor Name', 'Session Date', 'Session Type', 'Case ID', 'Status'
            ])

            session_rows = sessions.select_related('assigned_counselor__user').prefetch_related('participants')
            for s in session_rows:
                student_name = s.student.user.get_full_name() or s.student.user.username
                counselor = s.assigned_counselor.user.get_full_name() if s.assigned_counselor else ''
                session_date = s.scheduled_date or s.created_at.date()
                session_type = 'Group' if s.participants.exists() else 'Individual'
                case_obj = getattr(s, 'case', None)
                case_id = case_obj.id if case_obj else ''

                ws_sessions.append([
                    s.id,
                    s.student.student_number or '',
                    student_name,
                    s.student.course or '',
                    s.student.year_level or '',
                    counselor,
                    str(session_date),
                    session_type,
                    case_id,
                    s.get_status_display() if hasattr(s, 'get_status_display') else s.status,
                ])
            style_sheet(ws_sessions)

            # ---------- Sheet 3: Case Records ----------
            ws_cases = wb.create_sheet('Case Records')
            ws_cases.append([
                'Case ID', 'Student ID', 'Student Name', 'Case Type', 'Description',
                'Date Reported', 'Counselor Assigned', 'Case Status'
            ])

            status_map = {
                'pending': 'Pending',
                'in_progress': 'In Progress',
                'completed': 'Resolved',
                'cancelled': 'Cancelled',
            }
            for c in cases:
                student_name = c.student.user.get_full_name() or c.student.user.username
                counselor_name = c.counselor.user.get_full_name() if c.counselor else ''
                ws_cases.append([
                    c.id,
                    c.student.student_number or '',
                    student_name,
                    c.title,
                    c.description,
                    c.created_at.strftime('%Y-%m-%d'),
                    counselor_name,
                    status_map.get(c.status, c.status),
                ])
            style_sheet(ws_cases)

            # ---------- Sheet 4: Students Assisted ----------
            ws_students = wb.create_sheet('Students Assisted')
            ws_students.append([
                'Student ID', 'Student Name', 'Course', 'Year Level',
                'Contact Number', 'Email Address'
            ])

            student_profiles = Profile.objects.filter(user_type='student').select_related('user')
            for sp in student_profiles:
                student_name = sp.user.get_full_name() or sp.user.username
                ws_students.append([
                    sp.student_number or '',
                    student_name,
                    sp.course or '',
                    sp.year_level or '',
                    'N/A',
                    sp.user.email or '',
                ])
            style_sheet(ws_students)

            # ---------- Sheet 5: Monthly Counseling Statistics ----------
            ws_monthly = wb.create_sheet('Monthly Counseling Stats')
            ws_monthly.append(['Month', 'Total Counseling Sessions'])

            monthly_data = (
                GuidanceSession.objects
                .annotate(month=TruncMonth('created_at'))
                .values('month')
                .annotate(total=Count('id'))
                .order_by('month')
            )
            for row in monthly_data:
                label = row['month'].strftime('%B %Y') if row['month'] else ''
                ws_monthly.append([label, row['total']])
            style_sheet(ws_monthly)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            resp = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            resp['Content-Disposition'] = 'attachment; filename="CVSU_Guidance_Office_Report.xlsx"'
            return resp
        elif format == 'pdf':
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.units import cm
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
                from reportlab.platypus import (
                    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                    HRFlowable, Image as RLImage
                )
                from django.db.models.functions import TruncMonth
                from django.utils import timezone
                import io, os

                now = timezone.now()
                current_year = now.year
                counselor_name = request.user.get_full_name() or request.user.username

                # --- Collect statistics ---
                total_sessions = GuidanceSession.objects.count()
                total_cases = Case.objects.filter(is_active=True).count()
                total_students = Profile.objects.filter(user_type='student').count()
                completed_sessions = GuidanceSession.objects.filter(status='completed').count()
                completion_rate = int((completed_sessions / total_sessions * 100)) if total_sessions > 0 else 0

                pending_cases = Case.objects.filter(status='pending', is_active=True).count()
                in_progress_cases = Case.objects.filter(status='in_progress', is_active=True).count()
                completed_cases = Case.objects.filter(status='completed', is_active=True).count()
                cancelled_cases = Case.objects.filter(status='cancelled', is_active=True).count()

                # Sessions by month for current year
                month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                                'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                monthly_dict = {i: 0 for i in range(1, 13)}
                for row in (GuidanceSession.objects
                            .filter(created_at__year=current_year)
                            .annotate(m=TruncMonth('created_at'))
                            .values('m')
                            .annotate(cnt=Count('id'))):
                    monthly_dict[row['m'].month] = row['cnt']

                # Reason data
                reason_display_map = {
                    'academic': 'Academic Concerns',
                    'personal': 'Personal / Family Issues',
                    'career': 'Career Guidance',
                    'mental_health': 'Emotional / Mental Health Concerns',
                    'other': 'Behavioral / Other Concerns',
                }
                reason_data = list(
                    GuidanceSession.objects.values('reason')
                    .annotate(count=Count('reason'))
                    .order_by('-count')
                )

                reporting_period = f"Academic Year {current_year}-{current_year + 1}"

                # --- Build PDF ---
                buf = io.BytesIO()
                doc = SimpleDocTemplate(
                    buf, pagesize=A4,
                    rightMargin=2 * cm, leftMargin=2 * cm,
                    topMargin=1.5 * cm, bottomMargin=2 * cm,
                )
                styles = getSampleStyleSheet()

                s_center = ParagraphStyle('s_center', parent=styles['Normal'],
                                          fontSize=10, fontName='Helvetica',
                                          alignment=TA_CENTER, spaceAfter=2)
                s_subtitle = ParagraphStyle('s_subtitle', parent=styles['Normal'],
                                            fontSize=11, fontName='Helvetica-Bold',
                                            alignment=TA_CENTER, spaceAfter=2)
                s_big_title = ParagraphStyle('s_big_title', parent=styles['Normal'],
                                             fontSize=12, fontName='Helvetica-Bold',
                                             alignment=TA_CENTER, spaceAfter=4)
                s_section = ParagraphStyle('s_section', parent=styles['Normal'],
                                           fontSize=11, fontName='Helvetica-Bold',
                                           spaceBefore=10, spaceAfter=5,
                                           textColor=colors.HexColor('#1a3c5e'))
                s_body = ParagraphStyle('s_body', parent=styles['Normal'],
                                        fontSize=10, fontName='Helvetica',
                                        alignment=TA_JUSTIFY, leading=15, spaceAfter=6)
                s_footer = ParagraphStyle('s_footer', parent=styles['Normal'],
                                          fontSize=8, fontName='Helvetica-Oblique',
                                          alignment=TA_CENTER,
                                          textColor=colors.HexColor('#888888'))

                HDR_BG = colors.HexColor('#1a3c5e')
                ROW_A = colors.HexColor('#eef2f7')
                DIVIDER = colors.HexColor('#cccccc')
                WHITE = colors.white

                def table_style(has_total_row=False):
                    base = [
                        ('BACKGROUND', (0, 0), (-1, 0), HDR_BG),
                        ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 0), (-1, -1), 10),
                        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1 if not has_total_row else -2),
                         [ROW_A, WHITE]),
                        ('GRID', (0, 0), (-1, -1), 0.5, DIVIDER),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
                        ('TOPPADDING', (0, 0), (-1, -1), 7),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ]
                    if has_total_row:
                        base += [
                            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#dce7f0')),
                            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                        ]
                    return TableStyle(base)

                elems = []

                # ── HEADER ──────────────────────────────────────────────
                logo_candidates = [
                    os.path.join(str(settings.BASE_DIR), 'static', 'app', 'Images', 'Cavite_State_University_Logo.png'),
                    os.path.join(str(settings.BASE_DIR), 'staticfiles', 'app', 'Images', 'Cavite_State_University_Logo.png'),
                ]
                logo_path = next((p for p in logo_candidates if os.path.exists(p)), None)
                if logo_path:
                    logo = RLImage(logo_path, width=2.2 * cm, height=2.2 * cm)
                    logo.hAlign = 'CENTER'
                    elems.append(logo)
                    elems.append(Spacer(1, 0.25 * cm))

                elems.append(Paragraph('Republic of the Philippines', s_center))
                elems.append(Paragraph('<b>CAVITE STATE UNIVERSITY</b>', s_subtitle))
                elems.append(Paragraph('Bacoor City Campus', s_center))
                elems.append(Spacer(1, 0.2 * cm))
                elems.append(HRFlowable(width='100%', thickness=2, color=HDR_BG))
                elems.append(HRFlowable(width='100%', thickness=0.5, color=HDR_BG,
                                        spaceAfter=4))
                elems.append(Spacer(1, 0.2 * cm))
                elems.append(Paragraph('Guidance and Counseling Office', s_subtitle))
                elems.append(Paragraph('<u><b>NARRATIVE REPORT ON COUNSELING SERVICES</b></u>',
                                       s_big_title))
                elems.append(Spacer(1, 0.4 * cm))

                # ── I. REPORTING INFORMATION ─────────────────────────────
                elems.append(Paragraph('I.&nbsp;&nbsp;REPORTING INFORMATION', s_section))
                elems.append(HRFlowable(width='100%', thickness=0.5, color=DIVIDER))
                elems.append(Spacer(1, 0.2 * cm))
                info_table = Table([
                    ['Reporting Period:', reporting_period],
                    ['Date Generated:', now.strftime('%B %d, %Y')],
                    ['Prepared by:', counselor_name],
                    ['Office:', 'Guidance and Counseling Office'],
                ], colWidths=[4.5 * cm, 12 * cm])
                info_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                elems.append(info_table)

                # ── II. INTRODUCTION ─────────────────────────────────────
                elems.append(Paragraph('II.&nbsp;&nbsp;INTRODUCTION', s_section))
                elems.append(HRFlowable(width='100%', thickness=0.5, color=DIVIDER))
                elems.append(Spacer(1, 0.2 * cm))
                elems.append(Paragraph(
                    f"This Narrative Report presents a comprehensive summary of the counseling services, "
                    f"student support activities, and case management efforts conducted by the Guidance and "
                    f"Counseling Office of Cavite State University – Bacoor City Campus for the reporting "
                    f"period covering <b>{reporting_period}</b>. The report outlines the statistical data on "
                    f"guidance sessions conducted, cases handled, and students assisted, as recorded in the "
                    f"Case Tracking System. It also highlights the programs and activities implemented, "
                    f"challenges encountered, and recommendations aimed at further improving the delivery of "
                    f"guidance and counseling services to the student population.",
                    s_body))

                # ── III. SUMMARY OF STATISTICS ───────────────────────────
                elems.append(Paragraph('III.&nbsp;&nbsp;SUMMARY OF SYSTEM STATISTICS', s_section))
                elems.append(HRFlowable(width='100%', thickness=0.5, color=DIVIDER))
                elems.append(Spacer(1, 0.2 * cm))
                elems.append(Paragraph(
                    'The following table presents the key statistical data recorded in the Case Tracking '
                    'System for the reporting period:', s_body))
                stats_tbl = Table([
                    ['Statistical Indicator', 'Value'],
                    ['Total Counseling Sessions', str(total_sessions)],
                    ['Total Cases Recorded', str(total_cases)],
                    ['Total Students Assisted', str(total_students)],
                    ['Completed Sessions', str(completed_sessions)],
                    ['Case Completion Rate', f'{completion_rate}%'],
                ], colWidths=[11 * cm, 5.5 * cm])
                stats_tbl.setStyle(table_style())
                elems.append(stats_tbl)

                # ── IV. SESSIONS BY MONTH ────────────────────────────────
                elems.append(Paragraph('IV.&nbsp;&nbsp;COUNSELING SESSIONS BY MONTH', s_section))
                elems.append(HRFlowable(width='100%', thickness=0.5, color=DIVIDER))
                elems.append(Spacer(1, 0.2 * cm))
                peak_month = max(monthly_dict, key=monthly_dict.get)
                peak_count = monthly_dict[peak_month]
                peak_note = (f"The highest volume was recorded in "
                             f"<b>{month_names[peak_month - 1]}</b> with "
                             f"<b>{peak_count}</b> session(s). "
                             if peak_count > 0
                             else "No sessions have been recorded for the current year yet. ")
                elems.append(Paragraph(
                    f"The table below shows the monthly distribution of counseling sessions for "
                    f"<b>{current_year}</b>. {peak_note}"
                    f"This trend aids the office in resource planning and scheduling.", s_body))

                split_hdr = ['Month', 'Sessions', 'Month', 'Sessions']
                split_rows = [
                    [month_names[i], str(monthly_dict[i + 1]),
                     month_names[i + 6], str(monthly_dict[i + 7])]
                    for i in range(6)
                ]
                month_tbl = Table([split_hdr] + split_rows,
                                  colWidths=[3.5 * cm, 3 * cm, 3.5 * cm, 3 * cm])
                month_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), HDR_BG),
                    ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [ROW_A, WHITE]),
                    ('GRID', (0, 0), (-1, -1), 0.5, DIVIDER),
                    ('LINEAFTER', (1, 0), (1, -1), 1.5, colors.HexColor('#aaaaaa')),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                ])
                month_tbl.setStyle(month_style)
                elems.append(month_tbl)

                # ── V. CASES BY STATUS ───────────────────────────────────
                elems.append(Paragraph('V.&nbsp;&nbsp;CASES BY STATUS', s_section))
                elems.append(HRFlowable(width='100%', thickness=0.5, color=DIVIDER))
                elems.append(Spacer(1, 0.2 * cm))
                elems.append(Paragraph(
                    f"As of <b>{now.strftime('%B %d, %Y')}</b>, a total of <b>{total_cases}</b> "
                    f"case(s) have been recorded. The breakdown by current status is as follows:",
                    s_body))

                def pct(n): return f"{int(n / total_cases * 100) if total_cases else 0}%"

                status_tbl = Table([
                    ['Case Status', 'Number of Cases', 'Percentage'],
                    ['Pending', str(pending_cases), pct(pending_cases)],
                    ['In Progress', str(in_progress_cases), pct(in_progress_cases)],
                    ['Completed / Resolved', str(completed_cases), pct(completed_cases)],
                    ['Cancelled', str(cancelled_cases), pct(cancelled_cases)],
                    ['TOTAL', str(total_cases), '100%'],
                ], colWidths=[8 * cm, 5 * cm, 3.5 * cm])
                status_tbl.setStyle(table_style(has_total_row=True))
                elems.append(status_tbl)

                # ── VI. NATURE OF CASES ──────────────────────────────────
                elems.append(Paragraph(
                    'VI.&nbsp;&nbsp;NATURE OF CASES / COMMON STUDENT CONCERNS', s_section))
                elems.append(HRFlowable(width='100%', thickness=0.5, color=DIVIDER))
                elems.append(Spacer(1, 0.2 * cm))
                elems.append(Paragraph(
                    'The table below presents the distribution of counseling sessions by the nature '
                    'of concern or presenting problem:', s_body))

                total_r = sum(r['count'] for r in reason_data) or 1
                nature_rows = []
                shown = set()
                for r in reason_data:
                    label = reason_display_map.get(r['reason'],
                                                   r['reason'].replace('_', ' ').title())
                    shown.add(r['reason'])
                    nature_rows.append([label, str(r['count']),
                                        f"{int(r['count'] / total_r * 100)}%"])
                for key, label in reason_display_map.items():
                    if key not in shown:
                        nature_rows.append([label, '0', '0%'])

                nature_tbl = Table(
                    [['Nature of Concern', 'No. of Sessions', 'Percentage']] + nature_rows,
                    colWidths=[9 * cm, 4.5 * cm, 3 * cm])
                nature_tbl.setStyle(table_style())
                elems.append(nature_tbl)

                # ── VII. GUIDANCE PROGRAMS AND ACTIVITIES ────────────────
                elems.append(Paragraph(
                    'VII.&nbsp;&nbsp;GUIDANCE PROGRAMS AND ACTIVITIES', s_section))
                elems.append(HRFlowable(width='100%', thickness=0.5, color=DIVIDER))
                elems.append(Spacer(1, 0.2 * cm))
                elems.append(Paragraph(
                    f"The Guidance and Counseling Office conducted the following activities and "
                    f"programs during the {reporting_period}:", s_body))
                act_tbl = Table([
                    ['Program / Activity', 'Description'],
                    ['Student Orientation',
                     'Orientation for incoming and enrolled students on available guidance services.'],
                    ['Mental Health Awareness',
                     'Awareness sessions promoting student mental wellness and self-care practices.'],
                    ['Career Guidance Sessions',
                     'Individual and group sessions focused on career planning and academic direction.'],
                    ['Counseling Workshops',
                     'Workshops addressing coping strategies, stress management, and peer support.'],
                    ['Individual Counseling',
                     'One-on-one counseling for students with personal, academic, and behavioral concerns.'],
                ], colWidths=[5.5 * cm, 11 * cm])
                act_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), HDR_BG),
                    ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [ROW_A, WHITE]),
                    ('GRID', (0, 0), (-1, -1), 0.5, DIVIDER),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
                    ('TOPPADDING', (0, 0), (-1, -1), 7),
                ])
                act_tbl.setStyle(act_style)
                elems.append(act_tbl)

                # ── VIII. OBSERVATIONS ───────────────────────────────────
                elems.append(Paragraph('VIII.&nbsp;&nbsp;OBSERVATIONS', s_section))
                elems.append(HRFlowable(width='100%', thickness=0.5, color=DIVIDER))
                elems.append(Spacer(1, 0.2 * cm))
                top_concern = (reason_display_map.get(reason_data[0]['reason'],
                                                       reason_data[0]['reason'])
                               if reason_data else None)
                obs_parts = [
                    "Based on the data gathered for the reporting period, the following trends and "
                    "observations were noted by the Guidance and Counseling Office:"
                ]
                if total_sessions > 0:
                    obs_parts.append(
                        f" A total of <b>{total_sessions}</b> counseling session(s) were conducted, "
                        f"with <b>{completed_sessions}</b> session(s) successfully completed, yielding "
                        f"a completion rate of <b>{completion_rate}%</b>.")
                if top_concern:
                    obs_parts.append(
                        f" The most frequently cited concern among students was "
                        f"'<b>{top_concern}</b>', indicating a significant need for focused "
                        f"interventions in this area.")
                if in_progress_cases > 0:
                    obs_parts.append(
                        f" There are currently <b>{in_progress_cases}</b> case(s) in progress, "
                        f"requiring continued monitoring and follow-through by assigned counselors.")
                obs_parts.append(
                    " The data reflects a consistent demand for counseling services throughout the "
                    "academic year, underscoring the importance of an accessible and responsive "
                    "guidance program.")
                elems.append(Paragraph(''.join(obs_parts), s_body))

                # ── IX. CHALLENGES ───────────────────────────────────────
                elems.append(Paragraph('IX.&nbsp;&nbsp;CHALLENGES ENCOUNTERED', s_section))
                elems.append(HRFlowable(width='100%', thickness=0.5, color=DIVIDER))
                elems.append(Spacer(1, 0.2 * cm))
                elems.append(Paragraph(
                    "The Guidance and Counseling Office encountered the following challenges during "
                    "the reporting period: "
                    "(1) Limited student availability for follow-up sessions due to academic workload "
                    "and schedule conflicts; "
                    "(2) Persistent stigma among students regarding mental health services, leading to "
                    "reluctance in seeking counseling; "
                    "(3) Incomplete case documentation in some instances due to time constraints and "
                    "high case volume; "
                    "(4) Resource limitations in conducting large-scale group activities and awareness "
                    "campaigns; "
                    "(5) Difficulty tracking and following up on cases involving students with "
                    "irregular attendance.", s_body))

                # ── X. RECOMMENDATIONS ──────────────────────────────────
                elems.append(Paragraph('X.&nbsp;&nbsp;RECOMMENDATIONS', s_section))
                elems.append(HRFlowable(width='100%', thickness=0.5, color=DIVIDER))
                elems.append(Spacer(1, 0.2 * cm))
                elems.append(Paragraph(
                    "In light of the observations and challenges identified, the Guidance and "
                    "Counseling Office recommends the following: "
                    "(1) Strengthen awareness campaigns to reduce stigma and encourage students to "
                    "proactively seek counseling services; "
                    "(2) Enhance coordination with faculty and academic departments to facilitate "
                    "student referrals and follow-up engagements; "
                    "(3) Develop structured group counseling sessions to address common student "
                    "concerns more efficiently; "
                    "(4) Allocate additional resources and support staff to improve documentation "
                    "and case management practices; "
                    "(5) Establish a regular monitoring cycle for open and in-progress cases to "
                    "ensure timely resolution and proper case closure.", s_body))

                # ── XI. CONCLUSION ───────────────────────────────────────
                elems.append(Paragraph('XI.&nbsp;&nbsp;CONCLUSION', s_section))
                elems.append(HRFlowable(width='100%', thickness=0.5, color=DIVIDER))
                elems.append(Spacer(1, 0.2 * cm))
                elems.append(Paragraph(
                    f"The Guidance and Counseling Office of Cavite State University – Bacoor City "
                    f"Campus remains committed to providing holistic and responsive counseling services "
                    f"that support the academic, personal, and emotional well-being of every student. "
                    f"For the reporting period of <b>{reporting_period}</b>, the office recorded "
                    f"<b>{total_sessions}</b> counseling session(s) and handled <b>{total_cases}</b> "
                    f"case(s), assisting <b>{total_students}</b> student(s) in navigating their "
                    f"academic and personal challenges. The continuous improvement of guidance programs, "
                    f"the adoption of systematic case tracking through the Case Tracking System, and "
                    f"the dedication of the guidance staff collectively contribute to the realization "
                    f"of the university's mission of nurturing well-rounded and academically competent "
                    f"graduates.", s_body))

                # ── XII. SIGNATORIES ─────────────────────────────────────
                elems.append(Spacer(1, 0.8 * cm))
                elems.append(Paragraph('XII.&nbsp;&nbsp;SIGNATORIES', s_section))
                elems.append(HRFlowable(width='100%', thickness=0.5, color=DIVIDER))
                elems.append(Spacer(1, 0.6 * cm))

                s_sig_lbl = ParagraphStyle('s_sig_lbl', parent=styles['Normal'],
                                           fontSize=9, fontName='Helvetica',
                                           alignment=TA_CENTER)
                s_sig_name = ParagraphStyle('s_sig_name', parent=styles['Normal'],
                                            fontSize=10, fontName='Helvetica-Bold',
                                            alignment=TA_CENTER)
                s_sig_title = ParagraphStyle('s_sig_title', parent=styles['Normal'],
                                             fontSize=9, fontName='Helvetica',
                                             alignment=TA_CENTER)

                sig_row1 = Table([
                    [Paragraph('Prepared by:', s_sig_lbl), Paragraph('Reviewed by:', s_sig_lbl)],
                    [Paragraph('&nbsp;', s_sig_lbl), Paragraph('&nbsp;', s_sig_lbl)],
                    [Paragraph('&nbsp;', s_sig_lbl), Paragraph('&nbsp;', s_sig_lbl)],
                    [Paragraph(counselor_name, s_sig_name), Paragraph(' ', s_sig_name)],
                    [Paragraph('Guidance Counselor', s_sig_title),
                     Paragraph('Head, Guidance and Counseling Office', s_sig_title)],
                ], colWidths=[8.25 * cm, 8.25 * cm])
                sig_row1.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('LINEABOVE', (0, 3), (1, 3), 0.7, colors.black),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                elems.append(sig_row1)
                elems.append(Spacer(1, 1.2 * cm))

                sig_row2 = Table([
                    [Paragraph('Approved by:', s_sig_lbl)],
                    [Paragraph('&nbsp;', s_sig_lbl)],
                    [Paragraph('&nbsp;', s_sig_lbl)],
                    [Paragraph(' ', s_sig_name)],
                    [Paragraph('Campus Administrator / Campus Director', s_sig_title)],
                ], colWidths=[8.25 * cm])
                sig_row2.hAlign = 'CENTER'
                sig_row2.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('LINEABOVE', (0, 3), (0, 3), 0.7, colors.black),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                elems.append(sig_row2)

                # ── FOOTER ───────────────────────────────────────────────
                elems.append(Spacer(1, 0.6 * cm))
                elems.append(HRFlowable(width='100%', thickness=0.5, color=DIVIDER))
                elems.append(Paragraph(
                    f'<i>This report was automatically generated by the CVSU Case Tracking System '
                    f'on {now.strftime("%B %d, %Y at %I:%M %p")}.</i>',
                    s_footer))

                doc.build(elems)
                buf.seek(0)
                filename = f'CVSU_Guidance_Narrative_Report_{now.strftime("%Y%m%d")}.pdf'
                return FileResponse(buf, as_attachment=True, filename=filename)
            except Exception as e:
                import traceback
                return HttpResponse(
                    f'Error generating PDF: {e}\n\n{traceback.format_exc()}',
                    status=500, content_type='text/plain')
        else:
            return HttpResponse('Unsupported format', status=400)
    except Profile.DoesNotExist:
        return HttpResponse(status=403)


@login_required
def counselor_update_session_title(request, session_id):
    if request.method != 'POST':
        return redirect('counselor_session_detail', session_id=session_id)
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to perform this action.")
            return redirect('dashboard')
        session = get_object_or_404(GuidanceSession, id=session_id)
        title = (request.POST.get('title') or '').strip()
        if not title:
            messages.error(request, 'Title cannot be empty.')
            return redirect('counselor_session_detail', session_id=session.id)
        # Optionally ensure uniqueness by appending suffix if conflict (soft uniqueness)
        base = title
        suffix = 1
        while GuidanceSession.objects.filter(title=title).exclude(id=session.id).exists():
            suffix += 1
            title = f"{base} ({suffix})"
        session.title = title
        session.save()
        messages.success(request, 'Session title updated.')
        return redirect('counselor_session_detail', session_id=session.id)
    except Exception as e:
        messages.error(request, f'Error updating title: {e}')
        return redirect('counselor_session_detail', session_id=session_id)

@login_required
def check_new_notifications(request):
    """API endpoint to check for new notifications (for real-time polling)"""
    try:
        from app.utils import NotificationManager
        
        profile = Profile.objects.get(user=request.user)
        
        # Get the last notification ID from the request
        last_id = request.GET.get('last_id', 0)
        try:
            last_id = int(last_id)
        except (ValueError, TypeError):
            last_id = 0
        
        # Get new notifications since last_id
        new_notifications = Notification.objects.filter(
            recipient=profile,
            id__gt=last_id
        ).order_by('-created_at')[:5]
        
        # Get unread count
        unread_count = NotificationManager.get_unread_count(profile)
        
        # Serialize notifications
        notifications_data = []
        for notif in new_notifications:
            notifications_data.append({
                'id': notif.id,
                'title': notif.title,
                'message': notif.message,
                'link': notif.link or '#',
                'notification_type': notif.notification_type,
                'is_read': notif.is_read,
                'created_at': notif.created_at.isoformat()
            })
        
        return JsonResponse({
            'success': True,
            'new_notifications': notifications_data,
            'unread_count': unread_count
        })
    except Profile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Profile not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def counselor_create_session(request):
    """Create a new session (standalone or linked to case) with multiple participants"""
    if request.method == 'POST':
        try:
            from app.utils import NotificationManager
            from datetime import datetime
            
            profile = Profile.objects.get(user=request.user)
            if profile.user_type != 'counselor':
                return JsonResponse({'success': False, 'message': 'Not authorized'})
            
            # Get form data
            session_type = request.POST.get('session_type', 'standalone')
            title = request.POST.get('title')
            reason = request.POST.get('reason')
            student_id = request.POST.get('student')
            description = request.POST.get('description', '')
            scheduled_date = request.POST.get('scheduled_date')
            scheduled_time = request.POST.get('scheduled_time')
            location = request.POST.get('location', '')
            mode = request.POST.get('mode', 'in_person')
            auto_approve = request.POST.get('auto_approve') == 'on'
            send_notifications = request.POST.get('send_notifications') == 'on'
            assigned_counselor_id = request.POST.get('assigned_counselor')
            linked_case_id = request.POST.get('linked_case')
            case_id = request.POST.get('case_id')  # From button on case detail page
            
            # Get student
            student = Profile.objects.get(id=student_id)
            
            # Create session
            session = GuidanceSession.objects.create(
                title=title,
                reason=reason,
                student=student,
                student_number=student.student_number if hasattr(student, 'student_number') else 'N/A',
                description=description,
                scheduled_date=scheduled_date,
                scheduled_time=scheduled_time,
                location=location,
                status='approved' if auto_approve else 'pending',
                assigned_counselor=profile if not assigned_counselor_id else Profile.objects.get(id=assigned_counselor_id),
                preferred_counselor=profile
            )
            
            # Add participants
            participant_ids = request.POST.getlist('participants')
            if participant_ids:
                participants = Profile.objects.filter(id__in=participant_ids)
                session.participants.set(participants)
            
            # Link to case if specified
            if session_type == 'linked':
                if case_id:
                    case = Case.objects.get(id=case_id)
                    case.sessions.add(session)
                elif linked_case_id:
                    case = Case.objects.get(id=linked_case_id)
                    case.sessions.add(session)
            
            # Send notifications if enabled
            if send_notifications:
                # Notify primary student
                if auto_approve:
                    NotificationManager.notify_session_approved(session)
                    NotificationManager.notify_session_scheduled(session)
                else:
                    NotificationManager.create_notification(
                        'session_requested',
                        recipient=student,
                        sender=profile,
                        link=reverse('view_guidance_session_details', args=[session.id]),
                        student=student.user.get_full_name(),
                        reason=session.get_reason_display()
                    )
                
                # Notify all participants
                for participant in session.participants.all():
                    NotificationManager.create_notification(
                        'session_scheduled',
                        recipient=participant,
                        sender=profile,
                        link=reverse('view_guidance_session_details', args=[session.id]),
                        date=session.scheduled_date.strftime('%B %d, %Y'),
                        time=session.scheduled_time.strftime('%I:%M %p')
                    )
                
                # Notify assigned counselor if different from creator
                if assigned_counselor_id and assigned_counselor_id != str(profile.id):
                    assigned_counselor = Profile.objects.get(id=assigned_counselor_id)
                    NotificationManager.create_notification(
                        'case_assigned',
                        recipient=assigned_counselor,
                        link=reverse('counselor_session_detail', args=[session.id]),
                        title=f'Session: {title}'
                    )
            
            # Return success
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Session created successfully',
                    'session_id': session.id,
                    'redirect': reverse('counselor_session_detail', args=[session.id])
                })
            else:
                messages.success(request, 'Session created successfully!')
                return redirect('counselor_session_detail', session_id=session.id)
                
        except Profile.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Student not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def api_get_students(request):
    """API endpoint to get all students"""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            return JsonResponse({'error': 'Not authorized'}, status=403)
        
        students = Profile.objects.filter(user_type='student').select_related('user')
        data = [
            {
                'id': s.id,
                'name': s.user.get_full_name(),
                'student_number': s.student_number if hasattr(s, 'student_number') else 'N/A',
                'email': s.user.email
            }
            for s in students
        ]
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_get_profiles(request):
    """API endpoint to get all profiles (for participants)"""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            return JsonResponse({'error': 'Not authorized'}, status=403)
        
        profiles = Profile.objects.exclude(id=profile.id).select_related('user')
        data = [
            {
                'id': p.id,
                'name': p.user.get_full_name(),
                'type': p.get_user_type_display(),
                'email': p.user.email
            }
            for p in profiles
        ]
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_get_counselors(request):
    """API endpoint to get all counselors"""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            return JsonResponse({'error': 'Not authorized'}, status=403)
        
        counselors = Profile.objects.filter(user_type='counselor').select_related('user')
        data = [
            {
                'id': c.id,
                'name': c.user.get_full_name(),
                'email': c.user.email
            }
            for c in counselors
        ]
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_get_active_cases(request):
    """API endpoint to get all active cases"""
    try:
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            return JsonResponse({'error': 'Not authorized'}, status=403)
        
        cases = Case.objects.filter(
            is_active=True,
            status__in=['pending', 'in_progress']
        ).select_related('student__user')
        
        data = [
            {
                'id': c.id,
                'title': c.title,
                'student_name': c.student.user.get_full_name(),
                'status': c.get_status_display()
            }
            for c in cases
        ]
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# ============================================================================
# HEARING MANAGEMENT VIEWS
# ============================================================================

@login_required
def counselor_schedule_hearing(request):
    """Schedule a formal hearing for a case"""
    if request.method == 'POST':
        try:
            from app.utils import NotificationManager
            from app.models import Hearing, HearingAttendance
            
            profile = Profile.objects.get(user=request.user)
            if profile.user_type != 'counselor':
                return JsonResponse({'success': False, 'message': 'Not authorized'})
            
            # Get form data
            case_id = request.POST.get('case_id')
            hearing_type = request.POST.get('hearing_type')
            title = request.POST.get('title')
            scheduled_date = request.POST.get('scheduled_date')
            scheduled_time = request.POST.get('scheduled_time')
            estimated_duration = request.POST.get('estimated_duration', 60)
            location = request.POST.get('location')
            mode = request.POST.get('mode', 'in_person')
            meeting_link = request.POST.get('meeting_link', '')
            
            presiding_officer_name = request.POST.get('presiding_officer_name')
            panel_members_names = request.POST.get('panel_members_names', '')
            respondent_id = request.POST.get('respondent')
            complainant_id = request.POST.get('complainant')
            witness_ids = request.POST.getlist('witnesses')
            advisor_ids = request.POST.getlist('advisors')
            
            agenda = request.POST.get('agenda')
            charges = request.POST.get('charges', '')
            send_notifications = request.POST.get('send_notifications') == 'on'
            
            # Get case and participants
            case = Case.objects.get(id=case_id)
            respondent = Profile.objects.get(id=respondent_id)
            
            # For presiding officer, use the current counselor if name not found
            # Or you can create a text field to store the name
            presiding_officer = profile  # Default to current user
            
            # Create hearing
            hearing = Hearing.objects.create(
                case=case,
                title=title,
                hearing_type=hearing_type,
                scheduled_date=scheduled_date,
                scheduled_time=scheduled_time,
                estimated_duration=estimated_duration,
                location=location,
                mode=mode,
                meeting_link=meeting_link,
                presiding_officer=presiding_officer,
                respondent=respondent,
                agenda=agenda,
                charges=charges,
                created_by=profile
            )
            
            # Store presiding officer name and panel names in charges field as metadata
            # (You can add separate fields in model later if needed)
            metadata = {
                'presiding_officer_name': presiding_officer_name,
                'panel_members_names': panel_members_names.split('\n') if panel_members_names else []
            }
            hearing.charges = f"{charges}\n\n[METADATA: Presiding Officer: {presiding_officer_name}, Panel: {panel_members_names}]"
            hearing.save()
            
            # Note: Panel members as ManyToMany skipped since we're using text input
            # If you want to link to actual profiles, we can add that logic
            
            # Add complainant
            if complainant_id:
                hearing.complainant = Profile.objects.get(id=complainant_id)
                hearing.save()
            
            # Add witnesses
            if witness_ids:
                witnesses = Profile.objects.filter(id__in=witness_ids)
                hearing.witnesses.set(witnesses)
            
            # Add advisors
            if advisor_ids:
                advisors = Profile.objects.filter(id__in=advisor_ids)
                hearing.advisors.set(advisors)
            
            # Create attendance records
            HearingAttendance.objects.create(
                hearing=hearing,
                participant=presiding_officer,
                role='presiding_officer',
                status='confirmed'
            )
            
            for member in hearing.panel_members.all():
                HearingAttendance.objects.create(
                    hearing=hearing,
                    participant=member,
                    role='panel_member',
                    status='invited'
                )
            
            HearingAttendance.objects.create(
                hearing=hearing,
                participant=respondent,
                role='respondent',
                status='invited'
            )
            
            if hearing.complainant:
                HearingAttendance.objects.create(
                    hearing=hearing,
                    participant=hearing.complainant,
                    role='complainant',
                    status='invited'
                )
            
            for witness in hearing.witnesses.all():
                HearingAttendance.objects.create(
                    hearing=hearing,
                    participant=witness,
                    role='witness',
                    status='invited'
                )
            
            for advisor in hearing.advisors.all():
                HearingAttendance.objects.create(
                    hearing=hearing,
                    participant=advisor,
                    role='advisor',
                    status='invited'
                )
            
            # Send notifications if enabled
            if send_notifications:
                # Notify presiding officer (current user creating the hearing)
                NotificationManager.notify_hearing_scheduled(hearing, presiding_officer, 'presiding_officer')
                
                # Note: Panel members are text names, not profiles, so no notifications sent
                # If you want to notify them, you'll need to match names to profiles
                
                # Notify respondent
                NotificationManager.notify_hearing_scheduled(hearing, respondent, 'respondent')
                
                # Notify complainant
                if hearing.complainant:
                    NotificationManager.notify_hearing_scheduled(hearing, hearing.complainant, 'complainant')
                
                # Notify witnesses
                for witness in hearing.witnesses.all():
                    NotificationManager.notify_hearing_scheduled(hearing, witness, 'witness')
                
                # Notify advisors
                for advisor in hearing.advisors.all():
                    NotificationManager.notify_hearing_scheduled(hearing, advisor, 'advisor')
            
            # Return success
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Hearing {hearing.hearing_number} scheduled successfully',
                    'hearing_id': hearing.id,
                    'redirect': reverse('counselor_hearing_detail', args=[hearing.id])
                })
            else:
                messages.success(request, f'Hearing {hearing.hearing_number} scheduled successfully!')
                return redirect('counselor_hearing_detail', hearing_id=hearing.id)
                
        except Case.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Case not found'})
        except Profile.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Participant not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def counselor_hearing_detail(request, hearing_id):
    """View hearing details"""
    try:
        from app.models import Hearing
        
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('dashboard')
        
        hearing = get_object_or_404(Hearing, id=hearing_id)
        
        # Get all participants with their attendance status
        participants = hearing.get_all_participants()
        attendance_records = hearing.attendance.all()
        
        # Map attendance status to participants
        for participant in participants:
            attendance = attendance_records.filter(
                participant=participant['profile']
            ).first()
            participant['attendance'] = attendance
        
        # Get evidence
        evidence_list = hearing.evidence.all()
        
        context = {
            'profile': profile,
            'hearing': hearing,
            'participants': participants,
            'evidence_list': evidence_list,
            'can_edit': profile == hearing.presiding_officer or profile == hearing.created_by,
        }
        
        return render(request, 'app/counselor_hearing_detail.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Profile not found")
        return redirect('dashboard')


@login_required
def counselor_hearings_list(request):
    """List all hearings with filters"""
    try:
        from app.models import Hearing
        from django.core.paginator import Paginator
        
        profile = Profile.objects.get(user=request.user)
        if profile.user_type != 'counselor':
            messages.error(request, "You don't have permission to access this page.")
            return redirect('dashboard')
        
        # Get filter parameters
        status_filter = request.GET.get('status', 'all')
        
        # Base queryset
        hearings = Hearing.objects.all().select_related(
            'case', 'presiding_officer__user', 'respondent__user'
        ).order_by('-scheduled_date', '-scheduled_time')
        
        # Apply filters
        if status_filter != 'all':
            hearings = hearings.filter(status=status_filter)
        
        # Pagination
        page = request.GET.get('page', 1)
        paginator = Paginator(hearings, 20)
        hearings_page = paginator.get_page(page)
        
        # Count by status
        status_counts = {
            'all': Hearing.objects.count(),
            'scheduled': Hearing.objects.filter(status='scheduled').count(),
            'in_progress': Hearing.objects.filter(status='in_progress').count(),
            'completed': Hearing.objects.filter(status='completed').count(),
        }
        
        context = {
            'profile': profile,
            'hearings': hearings_page,
            'status_filter': status_filter,
            'status_counts': status_counts,
        }
        
        return render(request, 'app/counselor_hearings_list.html', context)
    except Profile.DoesNotExist:
        messages.error(request, "Profile not found")
        return redirect('dashboard')


@login_required
def counselor_hearing_upload_evidence(request, hearing_id):
    """Upload evidence for a hearing"""
    if request.method == 'POST':
        try:
            from app.models import Hearing, HearingEvidence
            from app.utils import NotificationManager
            
            profile = Profile.objects.get(user=request.user)
            hearing = get_object_or_404(Hearing, id=hearing_id)
            
            # Check if user is authorized to upload evidence
            participants = [p['profile'] for p in hearing.get_all_participants()]
            if profile not in participants:
                return JsonResponse({'success': False, 'message': 'Not authorized'})
            
            # Get form data
            evidence_type = request.POST.get('evidence_type')
            title = request.POST.get('title')
            description = request.POST.get('description', '')
            file = request.FILES.get('file')
            
            # Create evidence
            evidence = HearingEvidence.objects.create(
                hearing=hearing,
                submitted_by=profile,
                evidence_type=evidence_type,
                title=title,
                description=description,
                file=file
            )
            
            # Notify relevant parties
            relevant_parties = [
                hearing.presiding_officer,
                hearing.respondent
            ]
            if hearing.complainant:
                relevant_parties.append(hearing.complainant)
            
            NotificationManager.notify_hearing_evidence_uploaded(
                hearing, profile, relevant_parties
            )
            
            messages.success(request, 'Evidence uploaded successfully')
            return redirect('counselor_hearing_detail', hearing_id=hearing.id)
            
        except Exception as e:
            messages.error(request, f'Error uploading evidence: {str(e)}')
            return redirect('counselor_hearing_detail', hearing_id=hearing_id)
    
    return redirect('counselor_hearing_detail', hearing_id=hearing_id)


@login_required
def counselor_hearing_post_decision(request, hearing_id):
    """Post decision for a hearing"""
    if request.method == 'POST':
        try:
            from app.models import Hearing
            from app.utils import NotificationManager
            
            profile = Profile.objects.get(user=request.user)
            hearing = get_object_or_404(Hearing, id=hearing_id)
            
            # Only presiding officer can post decision
            if profile != hearing.presiding_officer:
                return JsonResponse({'success': False, 'message': 'Only presiding officer can post decision'})
            
            # Get form data
            verdict = request.POST.get('verdict')
            decision = request.POST.get('decision')
            sanctions = request.POST.get('sanctions', '[]')
            appeal_deadline = request.POST.get('appeal_deadline')
            
            # Update hearing
            hearing.verdict = verdict
            hearing.decision = decision
            hearing.sanctions = json.loads(sanctions) if sanctions else []
            if appeal_deadline:
                hearing.appeal_deadline = appeal_deadline
            hearing.status = 'completed'
            hearing.save()
            
            # Send notifications
            NotificationManager.notify_hearing_decision_posted(hearing)
            
            messages.success(request, 'Decision posted successfully')
            return redirect('counselor_hearing_detail', hearing_id=hearing.id)
            
        except Exception as e:
            messages.error(request, f'Error posting decision: {str(e)}')
            return redirect('counselor_hearing_detail', hearing_id=hearing_id)
    
    return redirect('counselor_hearing_detail', hearing_id=hearing_id)
